# Programa hecho por Duvelis Huiza modificada por Lic. Luis G.
from __future__ import annotations

import base64
from datetime import datetime, timedelta
import hmac
import json
import os
import re
import shutil
import sqlite3
import subprocess
import threading
import time
import sys
import unicodedata
from dataclasses import dataclass
from dataclasses import field
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlparse

import openpyxl

BASE_DIR = Path(__file__).resolve().parent
WEB_DIR = BASE_DIR / "webui"
STORAGE_DIR = BASE_DIR / "storage"

REPORTS_DIR = STORAGE_DIR / "reports"
REPORT_TEMPLATE = BASE_DIR / "Monitores_Electricos_20260308_1924.xlsx"
REPORT_FILENAME = "Reporte Voltguard.xlsx"
REPORT_PATH = REPORTS_DIR / REPORT_FILENAME
REPORT_META_PATH = REPORTS_DIR / "Reporte Voltguard.meta.json"

# Algunos navegadores/túneles pueden repetir el GET de descarga (reintentos, range, etc.).
# Usamos un rid (request id) enviado por la UI para que por 1 clic solo se cree 1 copia.
_recent_export_rids: dict[str, float] = {}

_report_lock = threading.Lock()
_report_reset_thread_started = False


def _dotenv_defines(key: str) -> bool:
    """Devuelve True si el archivo .env contiene una asignación para `key`.

    Los scrapers usan `load_dotenv(..., override=False)`, por lo que si WebUI
    inyecta una variable en `env`, esa variable le gana a .env.

    Esta función permite poner defaults desde WebUI sin pisar lo que el usuario
    configuró explícitamente en .env.
    """

    try:
        env_path = BASE_DIR / ".env"
        if not env_path.exists():
            return False
        for raw in env_path.read_text(encoding="utf-8", errors="ignore").splitlines():
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            if line.startswith("export "):
                line = line[len("export ") :].lstrip()
            if "=" not in line:
                continue
            k = line.split("=", 1)[0].strip()
            if k == key:
                return True
    except Exception:
        return False
    return False


PROVIDERS: dict[str, dict[str, str]] = {
    "growatt": {
        "label": "Growatt",
        "script": "growatt_scrape_dashboard.py",
        "log": "webui-growatt.log",
    },
    "shinemonitor": {
        "label": "ShineMonitor",
        "script": "shinemonitor_scrape_voltage.py",
        "log": "webui-shinemonitor.log",
    },
    "values": {
        "label": "Values",
        "script": "values_scrape_voltage.py",
        "log": "webui-values.log",
    },
}


@dataclass
class Job:
    provider: str
    started_at: float
    log_path: Path
    baseline_max_id: dict[str, int] | None
    pid: int | None = None
    exit_code: int | None = None
    running: bool = False
    stop_requested: bool = False
    stop_requested_at: float | None = None
    # Estado para métricas de red (cuadrícula / plantas-hora)
    log_parse_pos: int = 0
    seen_plant_events: set[str] = field(default_factory=set)
    grid_parse_pos: int = 0
    grid_targets: list[str] = field(default_factory=list)
    grid_status: dict[str, str] = field(default_factory=dict)
    grid_current: str | None = None


_jobs_lock = threading.Lock()
_jobs: dict[str, Job] = {}
_procs: dict[str, subprocess.Popen[str]] = {}
_log_base_pos: dict[str, int] = {}


def _other_running_providers(current: str | None = None) -> list[str]:
    with _jobs_lock:
        running = [k for k, j in _jobs.items() if j.running and (current is None or k != current)]
    return running


def _json_response(handler: BaseHTTPRequestHandler, status: int, payload: Any) -> None:
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(data)))
    handler.end_headers()
    handler.wfile.write(data)


def _read_body_json(handler: BaseHTTPRequestHandler) -> dict[str, Any]:
    length = int(handler.headers.get("Content-Length") or "0")
    raw = handler.rfile.read(length) if length else b"{}"
    try:
        return json.loads(raw.decode("utf-8"))
    except Exception:
        return {}


def _ensure_storage() -> None:
    STORAGE_DIR.mkdir(parents=True, exist_ok=True)


def _ensure_reports_dir() -> None:
    _ensure_storage()
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)


def _norm_key(s: str) -> str:
    s = (s or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _canonical_slot(slot: str) -> str:
    s = _norm_key(slot)
    if s in ("medianoche", "media noche", "midnight"):
        return "medianoche"
    if s in ("manana", "mañana", "morning"):
        return "manana"
    if s in ("mediodia", "medio dia", "medio dia ", "noon", "medio"):
        return "mediodia"
    if s in ("tarde", "afternoon"):
        return "tarde"
    return ""


def _safe_report_path(file_name: str) -> Path | None:
    """Resuelve un archivo dentro de storage/reports de forma segura.

    Solo permite nombres base (sin carpetas) y extensión .xlsx.
    """

    try:
        name = str(file_name or "").strip()
    except Exception:
        return None
    if not name:
        return None
    if name != Path(name).name:
        return None
    if not name.lower().endswith(".xlsx"):
        return None

    _ensure_reports_dir()
    base = REPORTS_DIR.resolve()
    p = (REPORTS_DIR / name).resolve()
    try:
        if base not in p.parents:
            return None
    except Exception:
        return None
    return p


def _report_history_list() -> list[dict[str, Any]]:
    _ensure_reports_dir()
    items: list[dict[str, Any]] = []
    try:
        for p in REPORTS_DIR.glob("*.xlsx"):
            # El historial debe representar descargas/exportaciones (copias timestamped),
            # no el reporte persistente "actual".
            if p.name == REPORT_FILENAME:
                continue
            try:
                st = p.stat()
                items.append(
                    {
                        "name": p.name,
                        "size": int(st.st_size),
                        "mtime": float(st.st_mtime),
                        "is_current": False,
                    }
                )
            except Exception:
                continue
    except Exception:
        return []
    items.sort(key=lambda x: (-(x.get("mtime") or 0.0), _norm_key(x.get("name") or "")))
    return items


def _report_history_copy_from_current(*, slot: str) -> Path | None:
    """Crea una copia timestamped del reporte persistente en storage/reports."""

    try:
        _ensure_reports_dir()
        if not REPORT_PATH.exists():
            return None
        slot_c = _canonical_slot(slot)
        slot_names_map = {
            "manana": "Mañana",
            "mediodia": "Medio Dia",
            "tarde": "Tarde",
            "medianoche": "MediaNoche",
        }
        slot_name = slot_names_map.get(slot_c, "Reporte")
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_name = f"Reporte de Monitores_{slot_name}_{ts}.xlsx"
        out_path = REPORTS_DIR / out_name
        i = 2
        while out_path.exists():
            out_name = f"Reporte de Monitores_{slot_name}_{ts}_{i}.xlsx"
            out_path = REPORTS_DIR / out_name
            i += 1
        shutil.copy2(REPORT_PATH, out_path)
        
        # Copiar también a la carpeta de descargas del usuario local
        try:
            downloads_dir = Path.home() / "Downloads"
            if downloads_dir.exists() and downloads_dir.is_dir():
                shutil.copy2(REPORT_PATH, downloads_dir / out_name)
        except Exception:
            pass
            
        return out_path
    except Exception:
        return None


def _export_rid_seen(rid: str) -> bool:
    """True si el rid ya se procesó recientemente (ventana corta)."""

    rid = (rid or "").strip()
    if not rid:
        return False
    now = time.time()
    # limpieza de rids antiguos
    try:
        ttl = 10 * 60.0
        for k, t in list(_recent_export_rids.items()):
            if (now - float(t)) > ttl:
                _recent_export_rids.pop(k, None)
    except Exception:
        pass
    return rid in _recent_export_rids


def _export_rid_mark(rid: str) -> None:
    rid = (rid or "").strip()
    if not rid:
        return
    _recent_export_rids[rid] = time.time()


def _report_meta_load() -> dict[str, Any]:
    try:
        if REPORT_META_PATH.exists():
            return json.loads(REPORT_META_PATH.read_text(encoding="utf-8", errors="ignore") or "{}")
    except Exception:
        return {}
    return {}


def _report_meta_save(meta: dict[str, Any]) -> None:
    try:
        _ensure_reports_dir()
        REPORT_META_PATH.write_text(json.dumps(meta, ensure_ascii=False), encoding="utf-8")
    except Exception:
        return


def _report_status() -> dict[str, Any]:
    meta = _report_meta_load()
    exists = REPORT_PATH.exists()
    updated_at = None
    try:
        if exists:
            updated_at = REPORT_PATH.stat().st_mtime
    except Exception:
        updated_at = None
    return {
        "exists": bool(exists),
        "filename": REPORT_FILENAME,
        "updated_at": updated_at,
        "last_slot": meta.get("last_slot"),
        "cleared_at": meta.get("cleared_at"),
    }


def _load_values_name_to_table() -> dict[str, str]:
    db_name = _provider_db_name("values")
    dbs = {d["name"] for d in _list_sqlite_files()}
    if not db_name or db_name not in dbs:
        return {}
    out: dict[str, str] = {}
    try:
        conn = _sqlite_open_by_name(db_name)
        try:
            cols = [r[1] for r in conn.execute("PRAGMA table_info(meta_monitors)").fetchall()]
            cols_l = {c.lower(): c for c in cols}
            name_col = cols_l.get("name") or cols_l.get("monitor_name") or cols_l.get("label")
            table_col = cols_l.get("table_name") or cols_l.get("table")
            if name_col and table_col:
                rows = conn.execute(f"SELECT {name_col}, {table_col} FROM meta_monitors").fetchall()
                for n, t in rows:
                    if n and t:
                        out[_norm_key(str(n))] = str(t)
        finally:
            conn.close()
    except Exception:
        return {}
    return out


def _load_shinemonitor_name_to_plant() -> dict[str, str]:
    path = STORAGE_DIR / "shinemonitor-plants.json"
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8", errors="ignore") or "{}")
        plants = data.get("plants") or []
        out: dict[str, str] = {}
        if isinstance(plants, list):
            for p in plants:
                name = str((p or {}).get("name") or "").strip()
                pid = str((p or {}).get("plant_id") or "").strip()
                if name and pid:
                    out[_norm_key(name)] = pid
        return out
    except Exception:
        return {}


def _load_growatt_name_to_table() -> dict[str, str]:
    # Growatt no tiene meta; mapeamos por nombre aproximado en el nombre de la tabla.
    path = STORAGE_DIR / "growatt-plants.json"
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8", errors="ignore") or "{}")
        plants = data.get("dropdown_plants") or []
    except Exception:
        plants = []

    db_name = _provider_db_name("growatt")
    dbs = {d["name"] for d in _list_sqlite_files()}
    if not db_name or db_name not in dbs:
        return {}

    tables: list[str] = []
    try:
        conn = _sqlite_open_by_name(db_name)
        try:
            tables = [t for t in _sqlite_tables(conn) if t.startswith("g_")]
        finally:
            conn.close()
    except Exception:
        tables = []

    out: dict[str, str] = {}
    for p in (plants or []):
        try:
            name = str((p or {}).get("name") or "").strip()
        except Exception:
            name = ""
        if not name:
            continue
        nk = _norm_key(name)
        # best: tabla que contenga todas las palabras del nombre
        words = [w for w in nk.split() if w]
        best = None
        best_score = -1
        for t in tables:
            tn = _norm_key(t.replace("_", " "))
            score = sum(1 for w in words if w in tn)
            if score > best_score:
                best_score = score
                best = t
        if best and best_score >= max(1, len(words) // 2):
            out[nk] = best
    return out


def _sqlite_latest_row(conn: sqlite3.Connection, table: str) -> dict[str, Any] | None:
    try:
        cur = conn.execute(f'SELECT * FROM "{table}" ORDER BY id DESC LIMIT 1')
        row = cur.fetchone()
        if not row:
            return None
        cols = [d[0] for d in (cur.description or [])]
        return {cols[i]: row[i] for i in range(len(cols))}
    except Exception:
        return None


def _sqlite_latest_ok_row(conn: sqlite3.Connection, table: str) -> dict[str, Any] | None:
    """Devuelve la fila más reciente con status='OK' si existe esa columna.

    Si la tabla no tiene columna status o la query falla, vuelve al latest row normal.
    """

    try:
        cur = conn.execute(f"SELECT * FROM \"{table}\" WHERE status='OK' ORDER BY id DESC LIMIT 1")
        row = cur.fetchone()
        if not row:
            return None
        cols = [d[0] for d in (cur.description or [])]
        return {cols[i]: row[i] for i in range(len(cols))}
    except Exception:
        return _sqlite_latest_row(conn, table)


def _sheet_is_row_slot_layout(ws) -> bool:
    try:
        a1 = ws.cell(1, 1).value
        b1 = ws.cell(1, 2).value
        if not isinstance(a1, str) or not isinstance(b1, str):
            return False
        return _norm_key(a1) in ("nodo", "node") and _norm_key(b1) in ("hora", "hour")
    except Exception:
        return False


def _slot_label_excel(slot: str) -> str:
    slot = _canonical_slot(slot)
    if slot == "medianoche":
        return "Media Noche"
    if slot == "manana":
        return "Mañana"
    if slot == "mediodia":
        return "Medio Dia"
    if slot == "tarde":
        return "Tarde"
    return ""


def _style_report_sheet(ws) -> None:
    """Aplica estilo al reporte sin cambiar datos/columnas.

    Reglas solicitadas:
    - Solo encabezados y celdas de nombre (planta/monitor/dispositivo) en azul oscuro.
    - Todo el cuerpo en blanco.
    - Separadores (fila vacía entre bloques) en gris intermedio.
    """

    try:
        ws.sheet_view.showGridLines = False
    except Exception:
        pass

    max_row = int(getattr(ws, "max_row", 0) or 0)
    max_col = int(getattr(ws, "max_column", 0) or 0)
    if max_row < 1 or max_col < 1:
        return

    # Paleta
    c_blue = "12324A"  # azul oscuro para encabezados/nombres
    c_white = "FFFFFF"  # cuerpo blanco
    c_sep = "B8C0CC"  # gris intermedio separadores
    c_border = "D0D5DD"  # borde gris claro
    c_text = "1F2937"  # texto oscuro
    c_muted = "6B7280"  # texto gris
    c_text_on_blue = "E8EEFF"  # texto claro sobre azul

    fill_body = openpyxl.styles.PatternFill("solid", fgColor=c_white)
    fill_header = openpyxl.styles.PatternFill("solid", fgColor=c_blue)
    fill_name = openpyxl.styles.PatternFill("solid", fgColor=c_blue)
    fill_sep = openpyxl.styles.PatternFill("solid", fgColor=c_sep)

    font_header = openpyxl.styles.Font(bold=True, color=c_text_on_blue)
    font_text = openpyxl.styles.Font(bold=False, color=c_text)
    font_muted = openpyxl.styles.Font(bold=False, color=c_muted)
    font_name = openpyxl.styles.Font(bold=True, color=c_text_on_blue)

    align_center = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    align_left = openpyxl.styles.Alignment(horizontal="left", vertical="center")
    align_right = openpyxl.styles.Alignment(horizontal="right", vertical="center")

    thin = openpyxl.styles.Side(style="thin", color=c_border)
    med = openpyxl.styles.Side(style="medium", color=c_border)

    def border_all(*, top=thin, bottom=thin, left=thin, right=thin):
        return openpyxl.styles.Border(top=top, bottom=bottom, left=left, right=right)

    border_none = openpyxl.styles.Border()

    # Detectar columnas clave por header
    header_vals = []
    for c in range(1, max_col + 1):
        v = ws.cell(1, c).value
        header_vals.append(str(v).strip() if v is not None else "")

    def find_col(name: str) -> int | None:
        nn = _norm_key(name)
        for i, h in enumerate(header_vals, start=1):
            if _norm_key(h) == nn:
                return i
        return None

    col_hora = find_col("Hora")
    col_ts = find_col("Timestamp")
    col_plant = find_col("plant_name")
    col_device = find_col("device_name")
    col_monitor = find_col("monitor_name")
    name_cols = [c for c in (col_plant, col_device, col_monitor) if c]
    col_name = name_cols[0] if name_cols else None

    # Set background general + header
    try:
        ws.row_dimensions[1].height = 22
    except Exception:
        pass

    for c in range(1, max_col + 1):
        cell = ws.cell(1, c)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center
        cell.border = border_all(bottom=med)

    # Filas separadoras (vacías): renderizar como espacio
    sep_rows: set[int] = set()
    for r in range(2, max_row + 1):
        any_value = False
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            if isinstance(v, str) and not v.strip():
                continue
            any_value = True
            break
        if not any_value:
            sep_rows.add(r)

    # Body
    for r in range(2, max_row + 1):
        try:
            ws.row_dimensions[r].height = 8 if r in sep_rows else 18
        except Exception:
            pass
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            if r in sep_rows:
                cell.fill = fill_sep
                cell.border = border_none
                continue

            cell.border = border_all()

            # Nombres (planta/monitor/dispositivo) en azul oscuro
            if c in name_cols:
                cell.fill = fill_name
                cell.font = font_name
                cell.alignment = align_left
                continue

            # Resto del cuerpo: blanco
            cell.fill = fill_body
            if col_hora and c == col_hora:
                cell.font = font_muted
                cell.alignment = align_center
            elif col_ts and c == col_ts:
                cell.font = font_text
                cell.alignment = align_center
            else:
                v = cell.value
                if isinstance(v, (int, float)):
                    cell.font = font_text
                    cell.alignment = align_right
                else:
                    cell.font = font_text
                    cell.alignment = align_center

    # Bloques: reforzar borde superior/inferior detectando grupos por Hora
    if col_hora:
        def _slot_of(v: Any) -> str:
            return _canonical_slot(str(v)) if isinstance(v, str) else ""

        r = 2
        while r <= max_row:
            if r in sep_rows:
                r += 1
                continue
            if _slot_of(ws.cell(r, col_hora).value) != "medianoche":
                r += 1
                continue
            s1 = _slot_of(ws.cell(r + 1, col_hora).value) if r + 1 <= max_row else ""
            s2 = _slot_of(ws.cell(r + 2, col_hora).value) if r + 2 <= max_row else ""
            s3 = _slot_of(ws.cell(r + 3, col_hora).value) if r + 3 <= max_row else ""
            if (s1, s2, s3) != ("manana", "mediodia", "tarde"):
                r += 1
                continue

            start_r = r
            end_r = r + 3
            for c in range(1, max_col + 1):
                top_cell = ws.cell(start_r, c)
                bot_cell = ws.cell(end_r, c)
                try:
                    top_cell.border = border_all(top=med, bottom=top_cell.border.bottom, left=top_cell.border.left, right=top_cell.border.right)
                except Exception:
                    top_cell.border = border_all(top=med)
                try:
                    bot_cell.border = border_all(top=bot_cell.border.top, bottom=med, left=bot_cell.border.left, right=bot_cell.border.right)
                except Exception:
                    bot_cell.border = border_all(bottom=med)

            r = end_r + 1

    # Column widths aproximados (sin tocar contenido)
    # Nombre más ancho, Hora mediana, resto compacto.
    try:
        for c in range(1, max_col + 1):
            letter = openpyxl.utils.get_column_letter(c)
            w = 16
            if col_name and c == col_name:
                w = 34
            elif col_hora and c == col_hora:
                w = 14
            elif col_ts and c == col_ts:
                w = 22
            ws.column_dimensions[letter].width = w
    except Exception:
        pass


def _update_report_shinemonitor_sheet(*, ws, conn_sm: sqlite3.Connection, slot: str) -> None:
    """Genera/actualiza la hoja Shine Monitor desde la BD.

    Objetivo: que el Excel refleje TODOS los monitores presentes en la SQLite (aunque cambien),
    con 4 filas por monitor (Media Noche/Mañana/Medio Dia/Tarde) y las columnas solicitadas.

    Cuando se exporta un slot, solo se actualiza la fila de ese slot; las otras 3 filas se preservan.
    Si el último registro no tiene datos (status='NO_DATA'), se escribe 'NO_DATA' en Timestamp.
    """

    slot = _canonical_slot(slot)
    slot_label = _slot_label_excel(slot)
    if not slot or not slot_label:
        return

    # Columnas requeridas (en el orden pedido)
    headers = [
        "plant_name",
        "device_name",
        "Hora",
        "Timestamp",
        "Battery Voltage",
        "PV Voltage(V)",
        "Inverter Voltage(V)",
        "Batt Current(A)",
        "Charger Current(A)",
        "Charger Power(W)",
        "PLoad(W)",
        "PGrid(W)",
        "work state",
        "rated power(W)",
        "Grid Voltage(V)",
        "PInverter(W)",
        "Accumulated Sell Power(kWh)",
        "Accumulated Load Power(kWh)",
        "Accumulated Self_Use Power(kWh)",
        "charger work enable",
        "Accumulated PV Power(kWh)",
    ]

    # Excel header -> DB column
    excel_to_db = {
        _norm_key("Timestamp"): "Timestamp",
        _norm_key("Battery Voltage"): "Battery Voltage(V)",
        _norm_key("PV Voltage(V)"): "PV Voltage(V)",
        _norm_key("Inverter Voltage(V)"): "Inverter Voltage(V)",
        _norm_key("Batt Current(A)"): "Batt Current(A)",
        _norm_key("Charger Current(A)"): "Charger Current(A)",
        _norm_key("Charger Power(W)"): "Charger Power(W)",
        _norm_key("PLoad(W)"): "PLoad(W)",
        _norm_key("PGrid(W)"): "PGrid(W)",
        _norm_key("work state"): "work state",
        _norm_key("rated power(W)"): "rated power(W)",
        _norm_key("Grid Voltage(V)"): "Grid Voltage(V)",
        _norm_key("PInverter(W)"): "PInverter(W)",
        _norm_key("Accumulated Sell Power(kWh)"): "Accumulated Sell Power(kWh)",
        _norm_key("Accumulated Load Power(kWh)"): "Accumulated Load Power(kWh)",
        _norm_key("Accumulated Self_Use Power(kWh)"): "Accumulated Self_Use Power(kWh)",
        _norm_key("charger work enable"): "charger work enable",
        _norm_key("Accumulated PV Power(kWh)"): "Accumulated PV Power(kWh)",
    }

    hour_labels = ["Media Noche", "Mañana", "Medio Dia", "Tarde"]

    def _iter_existing_rows() -> dict[tuple[str, str, str], dict[str, Any]]:
        """Lee datos existentes para preservarlos: (plant, device, slot) -> {header_norm: value}."""
        out: dict[tuple[str, str, str], dict[str, Any]] = {}
        try:
            # validar header actual
            current = []
            for c in range(1, len(headers) + 1):
                v = ws.cell(1, c).value
                current.append(str(v).strip() if v is not None else "")
            if [_norm_key(x) for x in current] != [_norm_key(x) for x in headers]:
                return {}

            last_plant = ""
            last_device = ""
            for r in range(2, ws.max_row + 1):
                pv = ws.cell(r, 1).value
                dv = ws.cell(r, 2).value
                hv = ws.cell(r, 3).value
                if isinstance(pv, str) and pv.strip():
                    last_plant = pv.strip()
                if isinstance(dv, str) and dv.strip():
                    last_device = dv.strip()
                if not isinstance(hv, str) or not hv.strip():
                    continue
                s = _canonical_slot(hv)
                if not s:
                    continue
                key = (_norm_key(last_plant), _norm_key(last_device), s)
                rowd: dict[str, Any] = {}
                for c in range(1, len(headers) + 1):
                    hn = _norm_key(headers[c - 1])
                    rowd[hn] = ws.cell(r, c).value
                out[key] = rowd
        except Exception:
            return {}
        return out

    existing = _iter_existing_rows()

    # DB: plant_id -> plant_name
    pid_to_plant: dict[str, str] = {}
    try:
        for pid, pname in conn_sm.execute("SELECT plant_id, plant_name FROM meta_plants").fetchall():
            if pid is None or pname is None:
                continue
            pid_to_plant[str(pid)] = str(pname)
    except Exception:
        pid_to_plant = {}

    # DB: meta_devices (principal)
    devices: dict[tuple[str, str], dict[str, Any]] = {}
    try:
        rows = conn_sm.execute("SELECT plant_id, device_name, table_name FROM meta_devices").fetchall()
        for r in rows:
            if not r or r[1] is None or r[2] is None:
                continue
            pid = str(r[0]) if r[0] is not None else ""
            plant = pid_to_plant.get(pid, "")
            device = str(r[1]).strip()
            table = str(r[2]).strip()
            if not device or not table:
                continue
            k = (_norm_key(plant), _norm_key(device))
            d = devices.setdefault(k, {"plant": plant, "device": device, "tables": []})
            if table not in d["tables"]:
                d["tables"].append(table)
    except Exception:
        devices = {}

    # Fallback: tablas reales en sqlite_master que no estén en meta_devices.
    # Filtramos solo tablas de medición (deben tener columna 'Timestamp').
    def _table_has_timestamp(table: str) -> bool:
        try:
            cols = [r[1] for r in conn_sm.execute(f"PRAGMA table_info('{table}')").fetchall()]
            return "Timestamp" in cols
        except Exception:
            return False

    try:
        data_tables = [
            t
            for t in _sqlite_tables(conn_sm)
            if not t.startswith("meta_") and not t.startswith("sqlite_") and _table_has_timestamp(t)
        ]
    except Exception:
        data_tables = []
    known_tables = set()
    for d in devices.values():
        for t in (d.get("tables") or []):
            known_tables.add(str(t))

    def _derive_device_name(table_name: str) -> str:
        # Nodo_El_Socorro -> Nodo El Socorro
        s = str(table_name or "").strip().replace("_", " ")
        s = re.sub(r"\s+", " ", s).strip()
        return s

    for t in data_tables:
        if t in known_tables:
            continue
        dev = _derive_device_name(t)
        k = ("", _norm_key(dev))
        d = devices.setdefault(k, {"plant": "", "device": dev, "tables": []})
        if t not in d["tables"]:
            d["tables"].append(t)

    device_list = list(devices.values())
    device_list.sort(key=lambda x: (_norm_key(x.get("plant") or ""), _norm_key(x.get("device") or "")))

    # Limpiar hoja (valores + merges) para reconstruir la grilla ordenada
    try:
        for rng in list(getattr(ws, "merged_cells", []).ranges):
            try:
                ws.unmerge_cells(str(rng))
            except Exception:
                pass
    except Exception:
        pass
    try:
        if ws.max_row:
            ws.delete_rows(1, ws.max_row)
    except Exception:
        # fallback: si delete_rows falla, intentamos borrar valores
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                try:
                    ws.cell(r, c).value = None
                except Exception:
                    pass

    # Escribir header
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c)
        cell.value = h
        try:
            cell.font = openpyxl.styles.Font(bold=True)
        except Exception:
            pass
    ws.freeze_panes = "A2"

    # Helper: fila más reciente (por id) entre varias tablas (incluye NO_DATA)
    def _latest_any_row(tables: list[str]) -> dict[str, Any] | None:
        best = None
        best_id = -1
        for t in tables:
            rd = _sqlite_latest_row(conn_sm, t)
            if not rd:
                continue
            try:
                rid = int(rd.get("id") or 0)
            except Exception:
                rid = 0
            if rid > best_id:
                best_id = rid
                best = rd
        return best

    # Construir filas
    row = 2
    for d in device_list:
        plant = str(d.get("plant") or "").strip()
        device = str(d.get("device") or "").strip()
        tables = list(d.get("tables") or [])
        if not device or not tables:
            continue

        # Preservar datos existentes para los 4 slots
        for i, hl in enumerate(hour_labels):
            r = row + i
            # plant/device solo en la primera fila del bloque; se fusionan verticalmente
            if i == 0:
                ws.cell(r, 1).value = plant
                ws.cell(r, 2).value = device
            ws.cell(r, 3).value = hl

            slot_i = _canonical_slot(hl)
            is_after = False
            try:
                slots_order = ["medianoche", "manana", "mediodia", "tarde"]
                if slots_order.index(slot_i) > slots_order.index(slot):
                    is_after = True
            except ValueError:
                pass
            k = (_norm_key(plant), _norm_key(device), slot_i)
            prev = {} if is_after else (existing.get(k) or {})
            for c, h in enumerate(headers, start=1):
                hn = _norm_key(h)
                if hn in (_norm_key("plant_name"), _norm_key("device_name"), _norm_key("hora")):
                    continue
                if hn in prev:
                    ws.cell(r, c).value = prev.get(hn)
                elif is_after:
                    ws.cell(r, c).value = None

        # Merge plant/device over 4 rows
        try:
            ws.merge_cells(start_row=row, start_column=1, end_row=row + 3, end_column=1)
        except Exception:
            pass
        try:
            ws.merge_cells(start_row=row, start_column=2, end_row=row + 3, end_column=2)
        except Exception:
            pass

        # Actualizar solo el slot pedido con la fila más reciente de la BD
        best = _latest_any_row(tables)
        if best is not None:
            status = str(best.get("status") or "").strip()
            # buscar fila del slot en este bloque
            for i, hl in enumerate(hour_labels):
                if _canonical_slot(hl) != slot:
                    continue
                r = row + i
                for c, h in enumerate(headers, start=1):
                    hn = _norm_key(h)
                    if hn in (_norm_key("plant_name"), _norm_key("device_name"), _norm_key("hora")):
                        continue
                    db_col = excel_to_db.get(hn)
                    if not db_col:
                        continue
                    val = best.get(db_col)
                    if hn == _norm_key("timestamp"):
                        # si no hay timestamp, mostrar el status (ej. NO_DATA)
                        if val in (None, "", " ") and status:
                            val = status
                    ws.cell(r, c).value = val
                break

        row += 4

        # Fila separadora (espacio visual entre bloques)
        sep_r = row
        ws.cell(sep_r, 1)  # crea la fila aunque quede vacía
        try:
            ws.row_dimensions[sep_r].height = 8
        except Exception:
            pass
        row += 1


def _update_report_values_sheet(*, ws, conn_values: sqlite3.Connection, slot: str) -> None:
    """Genera/actualiza la hoja Values desde la BD.

    - Lista monitores desde meta_monitors (y fallback a tablas m_* reales).
    - Columnas: monitor_name, Hora, y luego columnas de medición (mapeadas por meta_columns.header_text).
    - 4 filas por monitor (Media Noche/Mañana/Medio Dia/Tarde).
    - Al exportar un slot, solo se actualiza la fila del slot; los otros 3 slots se preservan.
    - Si no hay datos (sin filas o sin timestamp), se escribe 'NO_DATA' en Timestamp.
    """

    slot = _canonical_slot(slot)
    slot_label = _slot_label_excel(slot)
    if not slot or not slot_label:
        return

    hour_labels = ["Media Noche", "Mañana", "Medio Dia", "Tarde"]

    # Tablas de medición típicas en Values: m_*
    def _values_measurement_tables() -> list[str]:
        try:
            tabs = [t for t in _sqlite_tables(conn_values) if t.startswith("m_")]
        except Exception:
            tabs = []
        return tabs

    # meta_monitors: monitor_name -> table_name
    monitors: list[tuple[str, str]] = []
    try:
        rows = conn_values.execute("SELECT monitor_name, table_name FROM meta_monitors").fetchall()
        for r in rows:
            if not r or r[0] is None or r[1] is None:
                continue
            name = str(r[0]).strip()
            t = str(r[1]).strip()
            if name and t:
                monitors.append((name, t))
    except Exception:
        monitors = []

    known_tables = {t for _, t in monitors}

    def _derive_monitor_name(table_name: str) -> str:
        # m_SFApure_Inv1_B2V600_xxx -> SFApure Inv1 B2V600 xxx
        s = str(table_name or "").strip()
        if s.startswith("m_"):
            s = s[2:]
        s = s.replace("_", " ")
        s = re.sub(r"\s+", " ", s).strip()
        return s

    for t in _values_measurement_tables():
        if t in known_tables:
            continue
        monitors.append((_derive_monitor_name(t), t))

    # Orden estable
    monitors.sort(key=lambda x: (_norm_key(x[0]), _norm_key(x[1])))

    # meta_columns: (table_name, column_name) -> header_text
    coltext_by_table_col: dict[tuple[str, str], str] = {}
    try:
        rows = conn_values.execute("SELECT table_name, header_text, column_name FROM meta_columns").fetchall()
        for r in rows:
            if not r or r[0] is None or r[1] is None or r[2] is None:
                continue
            coltext_by_table_col[(str(r[0]), str(r[2]))] = str(r[1])
    except Exception:
        coltext_by_table_col = {}

    def _table_columns_after_first3(table_name: str) -> list[str]:
        try:
            cols = [r[1] for r in conn_values.execute(f"PRAGMA table_info('{table_name}')").fetchall()]
            return [c for c in (cols[3:] if len(cols) > 3 else []) if c]
        except Exception:
            return []

    # Definir headers dinámicos (unión) usando el orden de columnas en la tabla
    metric_headers: list[str] = []
    seen: set[str] = set()
    for _, t in monitors:
        for col in _table_columns_after_first3(t):
            ht = coltext_by_table_col.get((t, col)) or col
            # normalizar "Marca de Tiempo" a "Timestamp"
            if _norm_key(str(ht)) in ("marca de tiempo", "marca tiempo", "timestamp", "time"):
                ht = "Timestamp"
            ht = str(ht).strip()
            if not ht:
                continue
            nk = _norm_key(ht)
            if nk in seen:
                continue
            seen.add(nk)
            metric_headers.append(ht)

    # Forzar Timestamp al inicio si existe, conservando el orden original.
    # No usar list.index dentro de sort: CPython vacía la lista temporalmente durante sort.
    pos = {_norm_key(h): i for i, h in enumerate(metric_headers)}
    metric_headers.sort(key=lambda h: (0 if _norm_key(h) == _norm_key("Timestamp") else 1, pos.get(_norm_key(h), 10**9)))

    headers = ["monitor_name", "Hora"] + metric_headers

    # Leer datos existentes para preservarlos (monitor, slot) -> {header_norm: value}
    def _read_existing() -> dict[tuple[str, str], dict[str, Any]]:
        out: dict[tuple[str, str], dict[str, Any]] = {}
        try:
            current = []
            for c in range(1, len(headers) + 1):
                v = ws.cell(1, c).value
                current.append(str(v).strip() if v is not None else "")
            if [_norm_key(x) for x in current] != [_norm_key(x) for x in headers]:
                return {}

            last_mon = ""
            for r in range(2, ws.max_row + 1):
                mv = ws.cell(r, 1).value
                hv = ws.cell(r, 2).value
                if isinstance(mv, str) and mv.strip():
                    last_mon = mv.strip()
                if not isinstance(hv, str) or not hv.strip() or not last_mon:
                    continue
                s = _canonical_slot(hv)
                if not s:
                    continue
                key = (_norm_key(last_mon), s)
                rowd: dict[str, Any] = {}
                for c in range(1, len(headers) + 1):
                    hn = _norm_key(headers[c - 1])
                    rowd[hn] = ws.cell(r, c).value
                out[key] = rowd
        except Exception:
            return {}
        return out

    existing = _read_existing()

    # Limpiar hoja y reconstruir
    try:
        for rng in list(getattr(ws, "merged_cells", []).ranges):
            try:
                ws.unmerge_cells(str(rng))
            except Exception:
                pass
    except Exception:
        pass
    try:
        if ws.max_row:
            ws.delete_rows(1, ws.max_row)
    except Exception:
        pass

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c)
        cell.value = h
        try:
            cell.font = openpyxl.styles.Font(bold=True)
        except Exception:
            pass
    ws.freeze_panes = "A2"

    # Preconstruir mapeo: table -> {header_norm -> column_name}
    table_map: dict[str, dict[str, str]] = {}
    for _, t in monitors:
        m: dict[str, str] = {}
        for col in _table_columns_after_first3(t):
            ht = coltext_by_table_col.get((t, col)) or col
            if _norm_key(str(ht)) in ("marca de tiempo", "marca tiempo", "timestamp", "time"):
                ht = "Timestamp"
            m[_norm_key(str(ht))] = col
        table_map[t] = m

    def _latest_row(table: str) -> dict[str, Any] | None:
        return _sqlite_latest_row(conn_values, table)

    row = 2
    for mon_name, t in monitors:
        if not mon_name or not t:
            continue

        # Escribir bloque de 4 filas
        for i, hl in enumerate(hour_labels):
            r = row + i
            if i == 0:
                ws.cell(r, 1).value = mon_name
            ws.cell(r, 2).value = hl

            slot_i = _canonical_slot(hl)
            is_after = False
            try:
                slots_order = ["medianoche", "manana", "mediodia", "tarde"]
                if slots_order.index(slot_i) > slots_order.index(slot):
                    is_after = True
            except ValueError:
                pass
            prev = {} if is_after else (existing.get((_norm_key(mon_name), slot_i)) or {})
            for c, h in enumerate(headers, start=1):
                hn = _norm_key(h)
                if hn in (_norm_key("monitor_name"), _norm_key("hora")):
                    continue
                if hn in prev:
                    ws.cell(r, c).value = prev.get(hn)
                elif is_after:
                    ws.cell(r, c).value = None

        try:
            ws.merge_cells(start_row=row, start_column=1, end_row=row + 3, end_column=1)
        except Exception:
            pass

        # Actualizar solo el slot pedido
        best = _latest_row(t)
        # buscar fila del slot
        target_r = None
        for i, hl in enumerate(hour_labels):
            if _canonical_slot(hl) == slot:
                target_r = row + i
                break
        if target_r is not None:
            if best is None:
                # sin filas -> NO_DATA
                for c, h in enumerate(headers, start=1):
                    if _norm_key(h) == _norm_key("Timestamp"):
                        ws.cell(target_r, c).value = "NO_DATA"
                        break
            else:
                colmap = table_map.get(t) or {}
                ts_written = False
                for c, h in enumerate(headers, start=1):
                    hn = _norm_key(h)
                    if hn in (_norm_key("monitor_name"), _norm_key("hora")):
                        continue
                    db_col = colmap.get(hn)
                    if not db_col:
                        continue
                    val = best.get(db_col) if isinstance(best, dict) else None
                    if hn == _norm_key("Timestamp"):
                        ts_written = True
                        if val in (None, "", " "):
                            val = "NO_DATA"
                    ws.cell(target_r, c).value = val
                if not ts_written:
                    # si no existe columna Timestamp por cualquier razón, no hacemos nada extra
                    pass

        row += 4

        # Fila separadora (espacio visual entre bloques)
        sep_r = row
        ws.cell(sep_r, 1)
        try:
            ws.row_dimensions[sep_r].height = 8
        except Exception:
            pass
        row += 1


def _update_report_growatt_sheet(*, ws, conn_growatt: sqlite3.Connection, slot: str) -> None:
    """Genera/actualiza la hoja Growhatt desde la BD de Growatt.

    - Fuente de verdad: tablas reales g_* dentro de la SQLite.
    - Columnas: plant_name, Hora, y luego TODAS las columnas encontradas (normalizando update_time -> Timestamp).
    - 4 filas por planta (Media Noche/Mañana/Medio Dia/Tarde).
    - Al exportar un slot, solo se actualiza la fila del slot; los otros 3 slots se preservan.
    - Si no hay datos (sin filas o sin timestamp), se escribe 'NO_DATA' en Timestamp.
    """

    slot = _canonical_slot(slot)
    slot_label = _slot_label_excel(slot)
    if not slot or not slot_label:
        return

    hour_labels = ["Media Noche", "Mañana", "Medio Dia", "Tarde"]

    def _growatt_tables() -> list[str]:
        try:
            return [t for t in _sqlite_tables(conn_growatt) if t.startswith("g_")]
        except Exception:
            return []

    tables = _growatt_tables()
    tables.sort(key=lambda t: _norm_key(t))

    # Mapeo opcional tabla -> nombre amigable desde growatt-plants.json (si existe)
    name_by_table: dict[str, str] = {}
    try:
        path = STORAGE_DIR / "growatt-plants.json"
        if path.exists():
            data = json.loads(path.read_text(encoding="utf-8", errors="ignore") or "{}")
            plants = data.get("dropdown_plants") or []
            growatt_map = _load_growatt_name_to_table()
            if isinstance(plants, list):
                for p in plants:
                    try:
                        name = str((p or {}).get("name") or "").strip()
                    except Exception:
                        name = ""
                    if not name:
                        continue
                    t = growatt_map.get(_norm_key(name))
                    if t:
                        name_by_table[str(t)] = name
    except Exception:
        name_by_table = {}

    def _derive_plant_name(table_name: str) -> str:
        # g_Nodo_Algo_SERIAL_hash -> "Nodo Algo SERIAL"
        s = str(table_name or "").strip()
        if s.startswith("g_"):
            s = s[2:]
        parts = [p for p in s.split("_") if p]
        if parts and re.fullmatch(r"[0-9a-fA-F]{8}", parts[-1] or ""):
            parts = parts[:-1]
        s = " ".join(parts)
        s = re.sub(r"\s+", " ", s).strip()
        return s

    plants: list[tuple[str, str]] = []
    for t in tables:
        pname = (name_by_table.get(t) or _derive_plant_name(t)).strip()
        if not pname:
            pname = t
        plants.append((pname, t))
    plants.sort(key=lambda x: (_norm_key(x[0]), _norm_key(x[1])))

    # Construir headers dinámicos: unión de columnas en orden (normalizando update_time -> Timestamp)
    def _table_cols(table_name: str) -> list[str]:
        try:
            return [r[1] for r in conn_growatt.execute(f"PRAGMA table_info('{table_name}')").fetchall() if r and r[1]]
        except Exception:
            return []

    metric_headers: list[str] = []
    seen: set[str] = set()
    for _, t in plants:
        for col in _table_cols(t):
            if _norm_key(col) == _norm_key("id"):
                continue
            ht = "Timestamp" if _norm_key(col) in (_norm_key("update_time"), _norm_key("timestamp")) else str(col)
            nk = _norm_key(ht)
            if not nk or nk in seen:
                continue
            seen.add(nk)
            metric_headers.append(ht)

    # Forzar Timestamp al inicio si existe
    pos = {_norm_key(h): i for i, h in enumerate(metric_headers)}
    metric_headers.sort(key=lambda h: (0 if _norm_key(h) == _norm_key("Timestamp") else 1, pos.get(_norm_key(h), 10**9)))

    headers = ["plant_name", "Hora"] + metric_headers

    # Leer datos existentes para preservarlos: (plant, slot) -> {header_norm: value}
    def _read_existing() -> dict[tuple[str, str], dict[str, Any]]:
        out: dict[tuple[str, str], dict[str, Any]] = {}
        try:
            current = []
            for c in range(1, len(headers) + 1):
                v = ws.cell(1, c).value
                current.append(str(v).strip() if v is not None else "")
            if [_norm_key(x) for x in current] != [_norm_key(x) for x in headers]:
                return {}

            last_plant = ""
            for r in range(2, ws.max_row + 1):
                pv = ws.cell(r, 1).value
                hv = ws.cell(r, 2).value
                if isinstance(pv, str) and pv.strip():
                    last_plant = pv.strip()
                if not isinstance(hv, str) or not hv.strip() or not last_plant:
                    continue
                s = _canonical_slot(hv)
                if not s:
                    continue
                key = (_norm_key(last_plant), s)
                rowd: dict[str, Any] = {}
                for c in range(1, len(headers) + 1):
                    hn = _norm_key(headers[c - 1])
                    rowd[hn] = ws.cell(r, c).value
                out[key] = rowd
        except Exception:
            return {}
        return out

    existing = _read_existing()

    # Limpiar hoja y reconstruir
    try:
        for rng in list(getattr(ws, "merged_cells", []).ranges):
            try:
                ws.unmerge_cells(str(rng))
            except Exception:
                pass
    except Exception:
        pass
    try:
        if ws.max_row:
            ws.delete_rows(1, ws.max_row)
    except Exception:
        pass

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c)
        cell.value = h
        try:
            cell.font = openpyxl.styles.Font(bold=True)
        except Exception:
            pass
    ws.freeze_panes = "A2"

    # Preconstruir mapeo: table -> {header_norm -> column_name}
    table_map: dict[str, dict[str, str]] = {}
    for _, t in plants:
        m: dict[str, str] = {}
        cols = _table_cols(t)
        # timestamp preferido por tabla
        time_col = None
        for cand in ("update_time", "Timestamp", "timestamp", "inserted_at"):
            if any(_norm_key(c) == _norm_key(cand) for c in cols):
                # usar el nombre real tal cual está en la tabla
                time_col = next((c for c in cols if _norm_key(c) == _norm_key(cand)), None)
                if time_col:
                    break
        if time_col:
            m[_norm_key("Timestamp")] = str(time_col)
        for col in cols:
            if _norm_key(col) == _norm_key("id"):
                continue
            if time_col and _norm_key(col) == _norm_key(time_col):
                continue
            m[_norm_key(str(col))] = str(col)
        table_map[t] = m

    def _latest_row(table: str) -> dict[str, Any] | None:
        return _sqlite_latest_row(conn_growatt, table)

    row = 2
    for plant_name, t in plants:
        if not plant_name or not t:
            continue

        # Bloque 4 filas
        for i, hl in enumerate(hour_labels):
            r = row + i
            if i == 0:
                ws.cell(r, 1).value = plant_name
            ws.cell(r, 2).value = hl

            slot_i = _canonical_slot(hl)
            is_after = False
            try:
                slots_order = ["medianoche", "manana", "mediodia", "tarde"]
                if slots_order.index(slot_i) > slots_order.index(slot):
                    is_after = True
            except ValueError:
                pass
            prev = {} if is_after else (existing.get((_norm_key(plant_name), slot_i)) or {})
            for c, h in enumerate(headers, start=1):
                hn = _norm_key(h)
                if hn in (_norm_key("plant_name"), _norm_key("hora")):
                    continue
                if hn in prev:
                    ws.cell(r, c).value = prev.get(hn)
                elif is_after:
                    ws.cell(r, c).value = None

        try:
            ws.merge_cells(start_row=row, start_column=1, end_row=row + 3, end_column=1)
        except Exception:
            pass

        # Actualizar solo el slot pedido
        target_r = None
        for i, hl in enumerate(hour_labels):
            if _canonical_slot(hl) == slot:
                target_r = row + i
                break
        best = _latest_row(t)
        if target_r is not None:
            if best is None:
                for c, h in enumerate(headers, start=1):
                    if _norm_key(h) == _norm_key("Timestamp"):
                        ws.cell(target_r, c).value = "NO_DATA"
                        break
            else:
                colmap = table_map.get(t) or {}
                ts_written = False
                for c, h in enumerate(headers, start=1):
                    hn = _norm_key(h)
                    if hn in (_norm_key("plant_name"), _norm_key("hora")):
                        continue
                    db_col = colmap.get(hn)
                    if not db_col:
                        continue
                    val = best.get(db_col) if isinstance(best, dict) else None
                    if hn == _norm_key("Timestamp"):
                        ts_written = True
                        if val in (None, "", " "):
                            val = "NO_DATA"
                    ws.cell(target_r, c).value = val
                if not ts_written:
                    pass

        row += 4

        # Fila separadora (espacio visual entre bloques)
        sep_r = row
        ws.cell(sep_r, 1)
        try:
            ws.row_dimensions[sep_r].height = 8
        except Exception:
            pass
        row += 1


def _pick_best_value(row: dict[str, Any], header: str, provider: str) -> Any:
    h = _norm_key(header)
    if not h:
        return None

    # Campos comunes fecha/hora
    if any(k in h for k in ("fecha", "hora", "timestamp", "tiempo", "marca de tiempo")):
        for k in ("captured_at", "inserted_at", "last_seen_at", "update_time", "time", "marca_de_tiempo"):
            if k in row and row.get(k) is not None:
                return row.get(k)

    # Voltaje
    if "volt" in h:
        pref = []
        if provider == "values":
            pref = ["Inverter_Voltage_V", "Inverter_Voltage", "Inverter Voltage(V)"]
        elif provider == "shinemonitor":
            pref = ["Inverter Voltage(V)", "Inverter_Voltage_V"]
        else:
            pref = ["voltage", "Voltage", "Inverter Voltage(V)"]
        for k in pref:
            if k in row and row.get(k) not in (None, ""):
                return row.get(k)
        for k, v in row.items():
            if "volt" in _norm_key(k):
                return v

    # Match directo por nombre de columna
    for k, v in row.items():
        if _norm_key(k) == h:
            return v
    return None


def _find_slot_spans(ws) -> dict[str, dict[str, Any]]:
    """Detecta spans de columnas por franja horaria.

    Retorna dict slot -> {slot_row, min_col, max_col, header_row}.
    """

    spans: dict[str, dict[str, Any]] = {}
    # tokens normalizados para detección por substring
    slot_tokens = {
        "medianoche": ["medianoche", "media noche"],
        "manana": ["manana", "mañana"],
        "mediodia": ["mediodia", "medio dia", "medio día"],
        "tarde": ["tarde"],
    }

    merged = list(ws.merged_cells.ranges)
    merged_by_cell = {}
    for r in merged:
        for rr in range(r.min_row, r.max_row + 1):
            for cc in range(r.min_col, r.max_col + 1):
                merged_by_cell[(rr, cc)] = r

    max_scan_row = min(ws.max_row, 120)
    for r in range(1, max_scan_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            nv = _norm_key(v)
            key = ""
            for k, toks in slot_tokens.items():
                if any(t in nv for t in toks):
                    key = k
                    break
            if not key:
                continue
            if key in spans:
                continue

            mr = merged_by_cell.get((r, c))
            if mr:
                min_c, max_c = mr.min_col, mr.max_col
            else:
                min_c, max_c = c, c

            # header row: primera fila debajo con 2+ strings dentro del span
            header_row = None
            for rr in range(r + 1, min(r + 6, ws.max_row) + 1):
                strings = 0
                for cc in range(min_c, max_c + 1):
                    vv = ws.cell(rr, cc).value
                    if isinstance(vv, str) and _norm_key(vv):
                        strings += 1
                if strings >= 2:
                    header_row = rr
                    break
            spans[key] = {"slot_row": r, "min_col": min_c, "max_col": max_c, "header_row": header_row}
    return spans


def _clear_report_workbook(wb) -> None:
    """Limpia valores de data manteniendo headers y nombres.

    Regla: detecta spans por franja horaria y borra celdas desde data_start hasta el final.
    """

    for sh in wb.sheetnames:
        ws = wb[sh]
        # Layout real de este reporte: columnas Nodo/Hora y data a la derecha.
        if _sheet_is_row_slot_layout(ws):
            for r in range(2, ws.max_row + 1):
                for c in range(3, ws.max_column + 1):
                    cell = ws.cell(r, c)
                    # No se puede asignar a celdas fusionadas (no top-left)
                    try:
                        if isinstance(cell, openpyxl.cell.cell.MergedCell):
                            continue
                    except Exception:
                        pass
                    cell.value = None
            continue

        # Fallback: layouts antiguos por spans (por si se reusa en otras plantillas)
        spans = _find_slot_spans(ws)
        if not spans:
            continue
        any_header = next((s for s in spans.values() if s.get("header_row")), None)
        if not any_header:
            continue
        hr = int(any_header["header_row"])
        name_col = None
        for c in range(1, ws.max_column + 1):
            v = ws.cell(hr, c).value
            if isinstance(v, str):
                nv = _norm_key(v)
                if "planta" in nv or "monitor" in nv or "nombre" in nv:
                    name_col = c
                    break
        for slot, sp in spans.items():
            if not sp.get("header_row"):
                continue
            data_start = int(sp["header_row"]) + 1
            for r in range(data_start, ws.max_row + 1):
                if name_col is not None:
                    nm = ws.cell(r, name_col).value
                    if nm is None or str(nm).strip() == "":
                        pass
                for c in range(int(sp["min_col"]), int(sp["max_col"]) + 1):
                    ws.cell(r, c).value = None


def _generate_or_update_report(*, slot: str) -> Path:
    slot = _canonical_slot(slot)
    if not slot:
        raise ValueError("slot inválido")

    if not REPORT_TEMPLATE.exists():
        raise FileNotFoundError("No existe la plantilla de Excel")

    _ensure_reports_dir()

    # Cargar/crear workbook persistido
    if REPORT_PATH.exists():
        wb = openpyxl.load_workbook(REPORT_PATH)
    else:
        wb = openpyxl.load_workbook(REPORT_TEMPLATE)
        # Al crear por primera vez, limpiamos data para partir de un layout vacío.
        _clear_report_workbook(wb)
        wb.save(REPORT_PATH)

    # Construir mapas de nombres (reservados para compatibilidad con layouts previos)
    values_map = _load_values_name_to_table()
    growatt_map = _load_growatt_name_to_table()

    db_values = _provider_db_name("values")
    db_growatt = _provider_db_name("growatt")
    db_sm = _provider_db_name("shinemonitor")
    dbs = {d["name"] for d in _list_sqlite_files()}

    conn_values = _sqlite_open_by_name(db_values) if (db_values and db_values in dbs) else None
    conn_growatt = _sqlite_open_by_name(db_growatt) if (db_growatt and db_growatt in dbs) else None
    conn_sm = _sqlite_open_by_name(db_sm) if (db_sm and db_sm in dbs) else None
    sm_plants_map = _shinemonitor_plants(conn_sm) if conn_sm else {}

    try:
        # ShineMonitor: llenar por layout real (filas por Hora)
        if conn_sm is not None and "Shine Monitor" in wb.sheetnames:
            _update_report_shinemonitor_sheet(ws=wb["Shine Monitor"], conn_sm=conn_sm, slot=slot)

        # Values: llenar basado en BD (meta_monitors/meta_columns)
        if conn_values is not None and "Values" in wb.sheetnames:
            _update_report_values_sheet(ws=wb["Values"], conn_values=conn_values, slot=slot)

        # Growatt: llenar basado en BD (tablas g_*)
        if conn_growatt is not None and "Growhatt" in wb.sheetnames:
            _update_report_growatt_sheet(ws=wb["Growhatt"], conn_growatt=conn_growatt, slot=slot)

        # Estilo (todas las hojas existentes del reporte)
        for sh_name in list(wb.sheetnames):
            try:
                _style_report_sheet(wb[sh_name])
            except Exception:
                continue

        wb.save(REPORT_PATH)
    finally:
        try:
            if conn_values:
                conn_values.close()
        except Exception:
            pass
        try:
            if conn_growatt:
                conn_growatt.close()
        except Exception:
            pass
        try:
            if conn_sm:
                conn_sm.close()
        except Exception:
            pass

    meta = _report_meta_load()
    meta["last_slot"] = slot
    meta["updated_at"] = time.time()
    _report_meta_save(meta)
    return REPORT_PATH


def _clear_report() -> None:
    if not REPORT_TEMPLATE.exists():
        raise FileNotFoundError("No existe la plantilla de Excel")
    _ensure_reports_dir()
    wb = openpyxl.load_workbook(REPORT_TEMPLATE)
    _clear_report_workbook(wb)
    wb.save(REPORT_PATH)
    meta = _report_meta_load()
    meta["cleared_at"] = time.time()
    meta["updated_at"] = time.time()
    meta["last_slot"] = None
    _report_meta_save(meta)


def _seconds_until_2359(now: float | None = None) -> float:
    now_ts = time.time() if now is None else float(now)
    dt = datetime.fromtimestamp(now_ts)
    target = datetime(dt.year, dt.month, dt.day, 23, 59, 0)
    if target <= dt:
        target = target + timedelta(days=1)
    return max(5.0, (target - dt).total_seconds())


def _report_daily_reset_loop() -> None:
    while True:
        try:
            time.sleep(_seconds_until_2359())
            with _report_lock:
                _clear_report()
        except Exception:
            time.sleep(60)


def _ensure_report_reset_thread() -> None:
    global _report_reset_thread_started
    if _report_reset_thread_started:
        return
    _report_reset_thread_started = True
    threading.Thread(target=_report_daily_reset_loop, daemon=True).start()


def _events_file(provider: str) -> Path:
    return STORAGE_DIR / f"plant-events-{provider}.jsonl"


def _append_plant_event(*, provider: str, plant_id: str, outcome: str, ts: float) -> None:
    """Guarda un evento por planta para graficar Plantas/Hora.

    outcome: 'ok' | 'fail'
    """

    if not plant_id:
        return
    if outcome not in ("ok", "fail"):
        return
    try:
        _ensure_storage()
        path = _events_file(provider)
        rec = {"ts": float(ts), "provider": provider, "plant_id": str(plant_id), "outcome": outcome}
        with path.open("a", encoding="utf-8") as f:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")
    except Exception:
        return


def _load_events_last_24h(provider: str, now: float) -> list[dict[str, Any]]:
    path = _events_file(provider)
    if not path.exists():
        return []
    out: list[dict[str, Any]] = []
    min_ts = now - 24 * 3600
    try:
        for raw in path.read_text(encoding="utf-8", errors="ignore").splitlines():
            raw = raw.strip()
            if not raw:
                continue
            try:
                rec = json.loads(raw)
            except Exception:
                continue
            outcome = rec.get("outcome", "")
            if outcome not in ("ok", "fail"):
                continue
            try:
                ts = float(rec.get("ts"))
            except Exception:
                continue
            if ts < min_ts or ts > now + 5:
                continue
            out.append(rec)
    except Exception:
        return []
    return out


_re_sm_plant = re.compile(r"\bPlant\s+(\d+)\b", re.IGNORECASE)
_re_sm_fail = re.compile(r"\b(NO_TABLE|NO_TAB|NO_DATA)\b", re.IGNORECASE)

_re_gw_plant = re.compile(r"\bPlanta\s+(\d+)\s*/\s*(\d+)\b", re.IGNORECASE)
_re_gw_ok = re.compile(r"\bOK:\s*SQLite:\s*insertado\b", re.IGNORECASE)
_re_gw_err = re.compile(r"\b(ERROR|EXCEPTION|Traceback)\b", re.IGNORECASE)

_re_vals_item = re.compile(r"^\[(\d+)\s*/\s*(\d+)\]", re.IGNORECASE)
_re_vals_ok = re.compile(r"Dato enviado correctamente a la Base de datos", re.IGNORECASE)
_re_vals_nodata = re.compile(r"\bNO DATA:", re.IGNORECASE)
_re_vals_warn_retry = re.compile(r"\bWARN:\s*fallo\s+intento\b|\breintento\b", re.IGNORECASE)
_re_vals_err = re.compile(r"\bERROR:\b", re.IGNORECASE)

_re_sm_item = re.compile(r"^\[(\d+)\s*/\s*(\d+)\]\s*Plant\s+(\d+)", re.IGNORECASE)
_re_sm_no_inv = re.compile(r"\bNO_INVERTER\b", re.IGNORECASE)


def _parse_shinemonitor_fail_plants_full(log_path: Path) -> set[str]:
    fails: set[str] = set()
    try:
        if not log_path.exists():
            return fails
        txt = log_path.read_text(encoding="utf-8", errors="ignore")
        current_plant: str | None = None
        for line in txt.splitlines():
            m = _re_sm_plant.search(line)
            if m:
                current_plant = m.group(1)
            if _re_sm_fail.search(line):
                if current_plant:
                    fails.add(current_plant)
    except Exception:
        return fails
    return fails


def _parse_shinemonitor_fail_plants_incremental(job: Job) -> set[str]:
    """Devuelve plant_ids marcadas como fallo (NO_TABLE/NO_TAB) desde el log.

    Lee incrementalmente desde job.log_parse_pos para no re-scanear todo el archivo.
    """

    fails: set[str] = set()
    try:
        path = job.log_path
        if not path.exists():
            return fails
        size = path.stat().st_size
        start = int(job.log_parse_pos or 0)
        if start < 0 or start > size:
            start = 0

        with path.open("r", encoding="utf-8", errors="ignore") as f:
            f.seek(start)
            chunk = f.read()
            job.log_parse_pos = f.tell()

        current_plant: str | None = None
        for line in chunk.splitlines():
            m = _re_sm_plant.search(line)
            if m:
                current_plant = m.group(1)

            if _re_sm_fail.search(line):
                if current_plant:
                    fails.add(current_plant)
    except Exception:
        return fails
    return fails


def _hour_bucket(ts: float) -> int:
    return int(ts // 3600) * 3600


def _network_load_series(provider: str) -> dict[str, Any]:
    now = time.time()
    events = _load_events_last_24h(provider, now)

    # 24 buckets: hora actual hacia atrás
    end_hour = _hour_bucket(now)
    hours = [end_hour - i * 3600 for i in reversed(range(24))]
    counts_ok: dict[int, int] = {h: 0 for h in hours}
    counts_fail: dict[int, int] = {h: 0 for h in hours}

    for rec in events:
        try:
            ts = float(rec["ts"])
        except Exception:
            continue
        h = _hour_bucket(ts)
        outcome = rec.get("outcome", "ok")
        if h in counts_ok:
            if outcome == "ok":
                counts_ok[h] += 1
            elif outcome == "fail":
                counts_fail[h] += 1

    total_ok = sum(counts_ok.values())
    total_fail = sum(counts_fail.values())

    series = [
        {
            "hour": h,
            "ok": int(counts_ok.get(h, 0)),
            "fail": int(counts_fail.get(h, 0)),
            "label": time.strftime("%H:%M", time.localtime(h)),
        }
        for h in hours
    ]
    return {
        "provider": provider,
        "series": series,
        "total_ok": total_ok,
        "total_fail": total_fail,
        "updated_at": now,
    }


def _provider_db_name(provider: str) -> str | None:
    if provider == "growatt":
        return "Voltage Growatt.sqlite"
    if provider == "shinemonitor":
        return "Voltage  Shinemonitor.sqlite"
    if provider == "values":
        return "Voltage  Values.sqlite"
    return None


def _provider_expected_tables(conn: sqlite3.Connection, provider: str) -> list[str]:
    if provider == "growatt":
        return [t for t in _sqlite_tables(conn) if t.startswith("g_")]

    if provider == "shinemonitor":
        try:
            rows = conn.execute("SELECT table_name FROM meta_devices ORDER BY table_name").fetchall()
            return [str(r[0]) for r in rows if r and r[0]]
        except Exception:
            return []

    if provider == "values":
        try:
            rows = conn.execute("SELECT table_name FROM meta_monitors ORDER BY table_name").fetchall()
            return [str(r[0]) for r in rows if r and r[0]]
        except Exception:
            return []

    return []


def _grid_targets_from_storage(provider: str) -> list[str]:
    """Targets para el Status Grid, independientemente de SQLite.

    Devuelve una lista estable de IDs (string) para pintar celdas en el dashboard.
    - growatt: indices 1..N (según growatt-plants.json)
    - values: indices 1..N (según values-monitors.json)
    - shinemonitor: plant_id (según shinemonitor-plants.json)
    """

    try:
        if provider == "growatt":
            path = STORAGE_DIR / "growatt-plants.json"
            if not path.exists():
                path = STORAGE_DIR / "growatt-dashboard.json"
                if not path.exists():
                    return []
            data = json.loads(path.read_text(encoding="utf-8", errors="ignore") or "{}")
            plants = data.get("dropdown_plants") or data.get("visited") or data.get("plants") or []
            n = len(plants) if isinstance(plants, list) else 0
            return [str(i) for i in range(1, n + 1)]

        if provider == "values":
            path = STORAGE_DIR / "values-monitors.json"
            if not path.exists():
                return []
            data = json.loads(path.read_text(encoding="utf-8", errors="ignore") or "{}")
            n = int(data.get("count") or 0)
            if n <= 0 and isinstance(data.get("monitors"), list):
                n = len(data.get("monitors") or [])
            return [str(i) for i in range(1, n + 1)]

        if provider == "shinemonitor":
            path = STORAGE_DIR / "shinemonitor-plants.json"
            if not path.exists():
                return []
            data = json.loads(path.read_text(encoding="utf-8", errors="ignore") or "{}")
            plants = data.get("plants") or []
            out: list[str] = []
            if isinstance(plants, list):
                for p in plants:
                    try:
                        pid = str((p or {}).get("plant_id") or "").strip()
                    except Exception:
                        pid = ""
                    if pid:
                        out.append(pid)
            return out
    except Exception:
        return []

    return []


def _grid_target_names_from_storage(provider: str) -> dict[str, str]:
    """Mapea target_id (string) -> nombre legible para el Status Grid."""

    out: dict[str, str] = {}
    try:
        if provider == "growatt":
            path = STORAGE_DIR / "growatt-plants.json"
            if not path.exists():
                path = STORAGE_DIR / "growatt-dashboard.json"
                if not path.exists():
                    return {}
            data = json.loads(path.read_text(encoding="utf-8", errors="ignore") or "{}")
            plants = data.get("dropdown_plants") or data.get("visited") or data.get("plants") or []
            if not isinstance(plants, list):
                return {}
            for i, p in enumerate(plants, start=1):
                try:
                    if isinstance(p, dict):
                        name = str(p.get("name") or "").strip()
                    else:
                        name = str(p).strip()
                except Exception:
                    name = ""
                if name:
                    out[str(i)] = name
            return out

        if provider == "values":
            path = STORAGE_DIR / "values-monitors.json"
            if not path.exists():
                return {}
            data = json.loads(path.read_text(encoding="utf-8", errors="ignore") or "{}")
            monitors = data.get("monitors") or []
            if not isinstance(monitors, list):
                return {}
            for i, m in enumerate(monitors, start=1):
                try:
                    name = str((m or {}).get("name") or "").strip()
                except Exception:
                    name = ""
                if name:
                    out[str(i)] = name
            return out

        if provider == "shinemonitor":
            path = STORAGE_DIR / "shinemonitor-plants.json"
            if not path.exists():
                return {}
            data = json.loads(path.read_text(encoding="utf-8", errors="ignore") or "{}")
            plants = data.get("plants") or []
            if not isinstance(plants, list):
                return {}
            for p in plants:
                try:
                    pid = str((p or {}).get("plant_id") or "").strip()
                    name = str((p or {}).get("name") or "").strip()
                except Exception:
                    pid, name = "", ""
                if pid and name:
                    out[pid] = name
            return out
    except Exception:
        return {}

    return {}


def _grid_set_status(job: Job, key: str, status: str) -> None:
    if not key:
        return
    prev = (job.grid_status.get(key) or "pending").strip()

    # No degradar un OK a estados peores.
    if prev == "ok" and status in ("pending", "retry", "fail"):
        return

    if status not in ("pending", "ok", "fail", "retry"):
        status = "pending"
    job.grid_status[key] = status


def _grid_switch_current(job: Job, new_key: str) -> None:
    """Al cambiar al siguiente target, cerrar el anterior si quedó en retry."""

    if not new_key:
        return

    if job.grid_current and job.grid_current != new_key:
        prev = (job.grid_status.get(job.grid_current) or "pending").strip()
        if prev == "retry":
            job.grid_status[job.grid_current] = "fail"

    # Si vuelve a empezar el mismo target y venía fallando, marcar retry.
    if job.grid_current == new_key:
        prev = (job.grid_status.get(new_key) or "pending").strip()
        if prev in ("fail", "retry"):
            _grid_set_status(job, new_key, "retry")

    job.grid_current = new_key
    job.grid_status.setdefault(new_key, "pending")


def _shinemonitor_plant_ok_since_baseline(
    *,
    conn: sqlite3.Connection,
    plant_id: str,
    baseline: dict[str, int] | None,
    plants_map: dict[str, list[str]],
) -> bool | None:
    """True si la planta insertó alguna fila OK (y con voltaje) desde baseline.

    Devuelve None si no hay datos para evaluar.
    """

    if not baseline or not plants_map:
        return None
    tables = plants_map.get(str(plant_id)) or []
    if not tables:
        return None
    for t in tables:
        base_id = int((baseline.get(t) if isinstance(baseline, dict) else 0) or 0)
        cur_id = _sqlite_max_ok_voltage_id(conn, t, "Inverter Voltage(V)")
        if cur_id > base_id:
            return True
    return False


def _shinemonitor_plant_status_from_db(
    *,
    conn: sqlite3.Connection,
    plant_id: str,
    plants_map: dict[str, list[str]],
) -> str:
    """Devuelve status por planta basado en la última fila real en SQLite.

    Reglas:
    - OK si alguna tabla del plant tiene última fila status='OK' y voltaje parseable.
    - FAIL si existe al menos una fila pero ninguna cumple OK.
    - PENDING si no hay filas en ninguna tabla.

    Esto evita falsos rojos por mensajes transitorios del UI/log.
    """

    tables = plants_map.get(str(plant_id)) or []
    if not tables:
        return "pending"

    any_row = False
    any_ok = False

    for t in tables:
        try:
            row = conn.execute(
                f'SELECT status, "Inverter Voltage(V)" FROM "{t}" ORDER BY id DESC LIMIT 1'
            ).fetchone()
        except Exception:
            continue

        if not row:
            continue

        any_row = True
        st = str(row[0] or "").strip().upper()
        volt = row[1]
        if st == "OK" and _parse_float(volt) is not None:
            any_ok = True
            break

    if any_ok:
        return "ok"
    if any_row:
        return "fail"
    return "pending"


def _shinemonitor_latest_event_status(
    *,
    conn: sqlite3.Connection,
    plant_id: str,
) -> str | None:
    """Último status registrado en plant_events para esa planta."""

    try:
        row = conn.execute(
            "SELECT status FROM plant_events WHERE plant_id=? ORDER BY id DESC LIMIT 1",
            (str(plant_id),),
        ).fetchone()
        if not row:
            return None
        st = str(row[0] or "").strip()
        return st or None
    except Exception:
        return None


def _update_job_live_grid(provider: str, job: Job) -> None:
    """Parsea incrementalmente el log y actualiza job.grid_status.

    Estados:
    - pending: gris (no iniciado o sin resultado)
    - retry: amarillo (reintento en curso)
    - ok: verde
    - fail: rojo
    """

    try:
        path = job.log_path
        if not path.exists():
            return
        size = path.stat().st_size
        start = int(job.grid_parse_pos or 0)
        if start < 0 or start > size:
            start = 0

        with path.open("r", encoding="utf-8", errors="ignore") as f:
            f.seek(start)
            chunk = f.read()
            job.grid_parse_pos = f.tell()

        if not chunk:
            return

        # ShineMonitor: para evaluar OK real necesitamos mapa plant->tablas.
        sm_plants_map: dict[str, list[str]] = {}
        sm_conn: sqlite3.Connection | None = None
        baseline = job.baseline_max_id if isinstance(job.baseline_max_id, dict) else None

        if provider == "shinemonitor" and baseline is not None:
            db_name = _provider_db_name("shinemonitor")
            dbs = {d["name"] for d in _list_sqlite_files()}
            if db_name and db_name in dbs:
                try:
                    sm_conn = _sqlite_open_by_name(db_name)
                    sm_plants_map = _shinemonitor_plants(sm_conn)
                except Exception:
                    sm_conn = None
                    sm_plants_map = {}

        try:
            for line in chunk.splitlines():
                line = line.rstrip("\n")

                if provider == "growatt":
                    m = _re_gw_plant.search(line)
                    if m:
                        idx = str(int(m.group(1)))
                        _grid_switch_current(job, idx)
                        continue

                    if _re_gw_ok.search(line) and job.grid_current:
                        _grid_set_status(job, job.grid_current, "ok")
                        continue

                    if _re_gw_err.search(line) and job.grid_current:
                        _grid_set_status(job, job.grid_current, "fail")
                        continue

                elif provider == "values":
                    m = _re_vals_item.search(line)
                    if m:
                        idx = str(int(m.group(1)))
                        _grid_switch_current(job, idx)
                        continue

                    # NO DATA debe evaluarse ANTES que OK, porque el scraper
                    # puede imprimir ambos (NO DATA + Dato enviado con 0 columnas).
                    if _re_vals_nodata.search(line) and job.grid_current:
                        _grid_set_status(job, job.grid_current, "fail")
                        continue

                    if _re_vals_ok.search(line) and job.grid_current:
                        # Solo marcar OK si no fue previamente marcado como fail
                        # (por NO DATA en este mismo monitor).
                        cur = (job.grid_status.get(job.grid_current) or "pending").strip()
                        if cur != "fail":
                            _grid_set_status(job, job.grid_current, "ok")
                        continue

                    if _re_vals_warn_retry.search(line) and job.grid_current:
                        _grid_set_status(job, job.grid_current, "retry")
                        continue

                    if _re_vals_err.search(line) and job.grid_current:
                        # Si el log ya está en ERROR para este monitor, lo dejamos como retry
                        # (si hubo WARN) o fail si no hay señal de reintento.
                        cur = (job.grid_status.get(job.grid_current) or "pending").strip()
                        if cur != "retry":
                            _grid_set_status(job, job.grid_current, "fail")
                        continue

                elif provider == "shinemonitor":
                    m = _re_sm_item.search(line)
                    if m:
                        plant_id = str(m.group(3))

                        # Al entrar en una nueva planta, evaluar la anterior contra SQLite.
                        prev = job.grid_current
                        if prev and prev != plant_id and sm_conn is not None:
                            try:
                                ok = _shinemonitor_plant_ok_since_baseline(
                                    conn=sm_conn,
                                    plant_id=prev,
                                    baseline=baseline,
                                    plants_map=sm_plants_map,
                                )
                                if ok is True:
                                    _grid_set_status(job, prev, "ok")
                                elif ok is False:
                                    # Si no insertó nada OK, marcar fail al cerrar la planta.
                                    if (job.grid_status.get(prev) or "pending") not in ("fail",):
                                        _grid_set_status(job, prev, "fail")
                            except Exception:
                                pass

                        _grid_switch_current(job, plant_id)
                        continue

                    if _re_sm_fail.search(line) and job.grid_current:
                        _grid_set_status(job, job.grid_current, "fail")
                        continue

                    if _re_sm_no_inv.search(line) and job.grid_current:
                        _grid_set_status(job, job.grid_current, "fail")
                        continue
        finally:
            if sm_conn is not None:
                try:
                    sm_conn.close()
                except Exception:
                    pass
    except Exception:
        return


def _shinemonitor_plants(conn: sqlite3.Connection) -> dict[str, list[str]]:
    """Devuelve {plant_id: [table_name,...]} desde meta_devices.

    La operatividad de ShineMonitor es por planta (como el progreso [i/total] en el log).
    """

    out: dict[str, list[str]] = {}
    try:
        rows = conn.execute("SELECT plant_id, table_name FROM meta_devices").fetchall()
        for r in rows:
            if not r or r[0] is None or r[1] is None:
                continue
            pid = str(r[0])
            t = str(r[1])
            if not t:
                continue
            out.setdefault(pid, []).append(t)
    except Exception:
        return {}
    return out


def _sqlite_max_id(conn: sqlite3.Connection, table: str) -> int:
    try:
        row = conn.execute(f'SELECT MAX(id) FROM "{table}"').fetchone()
        v = row[0] if row else 0
        return int(v or 0)
    except Exception:
        return 0


def _sqlite_max_ok_id(conn: sqlite3.Connection, table: str) -> int:
    """MAX(id) contando solo filas OK si existe la columna status.

    Si la tabla no tiene status (o falla la query), vuelve a MAX(id).
    """

    try:
        row = conn.execute(f"SELECT MAX(id) FROM \"{table}\" WHERE status='OK'").fetchone()
        v = row[0] if row else 0
        return int(v or 0)
    except Exception:
        return _sqlite_max_id(conn, table)


def _sqlite_max_ok_voltage_id(conn: sqlite3.Connection, table: str, voltage_col: str) -> int:
    """MAX(id) contando solo filas OK con voltaje no vacío.

    Si la tabla no tiene status/columna o falla la query, vuelve a MAX(id) o a MAX(id) OK.
    """

    try:
        q = (
            f'SELECT id FROM "{table}" '
            f"WHERE status='OK' AND \"{voltage_col}\" IS NOT NULL AND \"{voltage_col}\" != '' "
            "ORDER BY id DESC LIMIT 1"
        )
        row = conn.execute(q).fetchone()
        v = row[0] if row else 0
        return int(v or 0)
    except Exception:
        return _sqlite_max_ok_id(conn, table)


def _snapshot_baseline_max_id(provider: str) -> dict[str, int] | None:
    """Snapshot por corrida: max(id) por tabla esperada.

    Se usa para medir operatividad: si un target inserta filas nuevas, MAX(id) sube.
    """

    db_name = _provider_db_name(provider)
    if not db_name:
        return None
    dbs = {d["name"] for d in _list_sqlite_files()}
    if db_name not in dbs:
        return None

    try:
        conn = _sqlite_open_by_name(db_name)
        with conn:
            tables = _provider_expected_tables(conn, provider)
            if not tables:
                return None

            out: dict[str, int] = {}
            for t in tables:
                try:
                    if provider == "shinemonitor":
                        out[t] = _sqlite_max_ok_voltage_id(conn, t, "Inverter Voltage(V)")
                    elif provider == "values":
                        out[t] = _sqlite_max_ok_voltage_id(conn, t, "Inverter_Voltage_V")
                    else:
                        out[t] = _sqlite_max_id(conn, t)
                except Exception:
                    out[t] = 0
            return out
    except Exception:
        return None


def _start_job(provider: str) -> Job:
    if provider not in PROVIDERS:
        raise ValueError("provider inválido")

    _ensure_storage()
    meta = PROVIDERS[provider]

    log_path = STORAGE_DIR / meta["log"]
    # Reiniciar log por corrida
    try:
        log_path.write_text("", encoding="utf-8")
    except Exception:
        pass

    # Reiniciar base offset: para una corrida nueva siempre mostramos desde 0.
    with _jobs_lock:
        _log_base_pos[provider] = 0

    env = os.environ.copy()
    others_running = _other_running_providers(current=provider)
    # Defaults solo si NO están en el entorno ni en .env.
    if "HEADLESS" not in env and not _dotenv_defines("HEADLESS"):
        env["HEADLESS"] = "true"
    if "PYTHONUNBUFFERED" not in env and not _dotenv_defines("PYTHONUNBUFFERED"):
        env["PYTHONUNBUFFERED"] = "1"

    # Values: por defecto, usar TURBO para evitar el reset pesado por monitor.
    if provider == "values" and "VALUES_TURBO" not in env and not _dotenv_defines("VALUES_TURBO"):
        env["VALUES_TURBO"] = "1"

    # ShineMonitor: en WebUI (y especialmente con carga) conviene subir timeouts por defecto.
    if provider == "shinemonitor":
        if "SHINE_DEFAULT_TIMEOUT_MS" not in env and not _dotenv_defines("SHINE_DEFAULT_TIMEOUT_MS"):
            env["SHINE_DEFAULT_TIMEOUT_MS"] = "60000" if others_running else "45000"
        if "SHINE_NAV_TIMEOUT_MS" not in env and not _dotenv_defines("SHINE_NAV_TIMEOUT_MS"):
            env["SHINE_NAV_TIMEOUT_MS"] = "120000" if others_running else "90000"

    # Values: permitir subir timeouts cuando hay concurrencia.
    if provider == "values":
        if "VALUES_DEFAULT_TIMEOUT_MS" not in env and not _dotenv_defines("VALUES_DEFAULT_TIMEOUT_MS"):
            env["VALUES_DEFAULT_TIMEOUT_MS"] = "45000"
        if "VALUES_NAV_TIMEOUT_MS" not in env and not _dotenv_defines("VALUES_NAV_TIMEOUT_MS"):
            env["VALUES_NAV_TIMEOUT_MS"] = "120000"

    # Usar el python del venv si existe, sino el del proceso actual.
    venv_py = BASE_DIR / ".venv" / "Scripts" / "python.exe"
    python_exe = str(venv_py) if venv_py.exists() else sys.executable

    script_path = BASE_DIR / meta["script"]
    if not script_path.exists():
        raise FileNotFoundError(f"No existe {script_path}")

    # Lanzar proceso y volcar salida al log.
    # Usamos -u para stdout sin buffer.
    f = log_path.open("a", encoding="utf-8", buffering=1)
    try:
        ts = time.strftime("%Y-%m-%d %H:%M:%S")
        keys = [
            "HEADLESS",
            "BROWSER",
            "PLANT_ID",
            "SHINE_DEFAULT_TIMEOUT_MS",
            "SHINE_NAV_TIMEOUT_MS",
            "VALUES_TURBO",
            "VALUES_DEFAULT_TIMEOUT_MS",
            "VALUES_NAV_TIMEOUT_MS",
            "VALUES_USE_DEVICE_LIST",
            "VALUES_LIMIT_MONITORS",
        ]
        f.write(f"[WEBUI] started_at={ts}\n")
        f.write(f"[WEBUI] provider={provider}\n")
        f.write(f"[WEBUI] python={python_exe}\n")
        f.write(f"[WEBUI] script={script_path.name}\n")
        shown = {k: env.get(k) for k in keys if env.get(k) is not None}
        if shown:
            f.write(f"[WEBUI] env={json.dumps(shown, ensure_ascii=False)}\n")
        f.write("[WEBUI] ---\n")
        f.flush()
    except Exception:
        pass

    baseline = _snapshot_baseline_max_id(provider)
    proc = subprocess.Popen(
        [python_exe, "-u", str(script_path)],
        cwd=str(BASE_DIR),
        env=env,
        stdout=f,
        stderr=subprocess.STDOUT,
        text=True,
    )
    job = Job(
        provider=provider,
        started_at=time.time(),
        log_path=log_path,
        baseline_max_id=baseline,
        pid=proc.pid,
        running=True,
    )

    # Inicializar targets del grid lo antes posible para que se vean celdas grises
    # incluso si SQLite todavía no está disponible.
    try:
        job.grid_targets = _grid_targets_from_storage(provider)
        for t in job.grid_targets:
            job.grid_status.setdefault(t, "pending")
    except Exception:
        pass

    def _waiter() -> None:
        code = proc.wait()
        try:
            f.flush()
        except Exception:
            pass
        try:
            f.close()
        except Exception:
            pass
        with _jobs_lock:
            j = _jobs.get(provider)
            if j:
                j.exit_code = code
                j.running = False
        with _jobs_lock:
            _procs.pop(provider, None)

        # Consolidar WAL al final para que el archivo .sqlite refleje datos/schema.
        try:
            db_name = _provider_db_name(provider)
            if db_name:
                _sqlite_checkpoint(db_name)
        except Exception:
            pass

        # Guardar último status grid y eventos de tendencia.
        try:
            _finalize_run(provider, job)
        except Exception:
            pass

    with _jobs_lock:
        _jobs[provider] = job
        _procs[provider] = proc

    threading.Thread(target=_waiter, daemon=True).start()
    return job


def _job_status(provider: str) -> dict[str, Any]:
    with _jobs_lock:
        j = _jobs.get(provider)
        p = _procs.get(provider)
    if not j:
        return {
            "provider": provider,
            "label": PROVIDERS.get(provider, {}).get("label", provider),
            "running": False,
            "started_at": None,
            "pid": None,
            "exit_code": None,
            "log": str((STORAGE_DIR / PROVIDERS.get(provider, {}).get("log", "")).name),
        }
    if p and j.running:
        code = p.poll()
        if code is not None:
            j.exit_code = code
            j.running = False
            with _jobs_lock:
                _procs.pop(provider, None)

    return {
        "provider": provider,
        "label": PROVIDERS[provider]["label"],
        "running": bool(j.running),
        "started_at": j.started_at,
        "pid": j.pid,
        "exit_code": j.exit_code,
        "log": j.log_path.name,
        "stop_requested": bool(j.stop_requested),
        "stop_requested_at": j.stop_requested_at,
    }


def _stop_job(provider: str) -> dict[str, Any]:
    with _jobs_lock:
        proc = _procs.get(provider)
        job = _jobs.get(provider)

    if not proc or not job or not job.running:
        return {"ok": False, "error": "no está ejecutándose", "status": _job_status(provider)}

    try:
        with _jobs_lock:
            j = _jobs.get(provider)
            if j:
                j.stop_requested = True
                j.stop_requested_at = time.time()

        proc.terminate()
        try:
            proc.wait(timeout=5)
        except Exception:
            proc.kill()
    except Exception as e:
        return {"ok": False, "error": f"no se pudo detener: {type(e).__name__}: {e}", "status": _job_status(provider)}

    with _jobs_lock:
        _procs.pop(provider, None)
        job.running = False
        # exit_code puede quedar None si se kill; lo dejamos así.

    return {"ok": True, "status": _job_status(provider)}


def _read_log(provider: str, pos: int) -> dict[str, Any]:
    meta = PROVIDERS.get(provider)
    if not meta:
        return {"text": "", "pos": pos}

    path = STORAGE_DIR / meta["log"]
    if not path.exists():
        return {"text": "", "pos": pos}

    with _jobs_lock:
        base_pos = int(_log_base_pos.get(provider, 0) or 0)

    if pos < base_pos:
        pos = base_pos

    with path.open("rb") as f:
        try:
            f.seek(max(0, pos))
        except Exception:
            f.seek(0)
        data = f.read(64 * 1024)
        new_pos = f.tell()

    text = data.decode("utf-8", errors="replace")
    return {"text": text, "pos": new_pos}


def _clear_log(provider: str) -> dict[str, Any]:
    """Marca el log como 'limpiado' sin truncar el archivo.

    En Windows, truncar el archivo mientras el proceso lo está escribiendo puede fallar.
    Usamos un offset base: los reads empiezan desde el tamaño actual del archivo.
    """

    meta = PROVIDERS.get(provider)
    if not meta:
        return {"ok": False, "error": "provider inválido"}

    path = STORAGE_DIR / meta["log"]
    try:
        size = path.stat().st_size if path.exists() else 0
    except Exception:
        size = 0

    with _jobs_lock:
        _log_base_pos[provider] = int(size)

    return {"ok": True, "provider": provider, "base_pos": int(size)}


def _list_sqlite_files() -> list[dict[str, Any]]:
    """Lista DBs sqlite en la raíz del proyecto.

    Nota: los scrapers guardan en archivos tipo "Voltage*.sqlite".
    """

    items: list[dict[str, Any]] = []
    for p in sorted(BASE_DIR.glob("*.sqlite"), key=lambda x: x.name.lower()):
        try:
            st = p.stat()
            items.append(
                {
                    "name": p.name,
                    "size": int(st.st_size),
                    "mtime": float(st.st_mtime),
                }
            )
        except Exception:
            items.append({"name": p.name, "size": None, "mtime": None})
    return items


def _sqlite_open_by_name(db_name: str) -> sqlite3.Connection:
    allowed = {d["name"] for d in _list_sqlite_files()}
    if db_name not in allowed:
        raise ValueError("db inválida")

    path = BASE_DIR / db_name
    if not path.exists():
        raise FileNotFoundError("db no existe")

    # Abrir en modo solo-lectura para no interferir con el proceso writer.
    # Usamos URI porque maneja espacios y permite mode=ro.
    uri = path.resolve().as_uri() + "?mode=ro"
    conn = sqlite3.connect(
        uri,
        uri=True,
        timeout=0.5,
        check_same_thread=False,
        isolation_level=None,  # autocommit: evita transacciones largas de lectura
    )
    try:
        conn.execute("PRAGMA query_only=ON")
    except Exception:
        pass
    try:
        conn.execute("PRAGMA busy_timeout=250")
    except Exception:
        pass
    try:
        conn.execute("PRAGMA read_uncommitted=1")
    except Exception:
        pass
    return conn


def _sqlite_open_rw_by_name(db_name: str) -> sqlite3.Connection:
    """Abre la DB en modo lectura-escritura para operaciones de escritura (DROP TABLE, etc.)."""

    allowed = {d["name"] for d in _list_sqlite_files()}
    if db_name not in allowed:
        raise ValueError("db inválida")

    path = BASE_DIR / db_name
    if not path.exists():
        raise FileNotFoundError("db no existe")

    conn = sqlite3.connect(
        str(path),
        timeout=10,
        check_same_thread=False,
    )
    try:
        conn.execute("PRAGMA busy_timeout=5000")
    except Exception:
        pass
    try:
        conn.execute("PRAGMA journal_mode=WAL")
    except Exception:
        pass
    return conn

def _sqlite_checkpoint(db_name: str) -> None:
    """Fuerza checkpoint de WAL para que datos/schema se escriban en el .sqlite.

    Esto evita casos donde el .sqlite queda en 4KB y todo está en -wal, lo cual
    puede verse como "sin tablas" en algunos visores.
    """

    path = (BASE_DIR / db_name)
    if not path.exists():
        return

    # Intentar varias veces por si hay un lector momentáneo.
    for _ in range(5):
        try:
            conn = sqlite3.connect(str(path), timeout=1.5, check_same_thread=False, isolation_level=None)
            try:
                row = conn.execute("PRAGMA wal_checkpoint(TRUNCATE)").fetchone()
                # Formato esperado: (busy, log, checkpointed)
                busy = int(row[0]) if row and row[0] is not None else 0
                if busy == 0:
                    return
            finally:
                try:
                    conn.close()
                except Exception:
                    pass
        except Exception:
            pass
        time.sleep(0.2)
    return


def _sqlite_tables(conn: sqlite3.Connection) -> list[str]:
    rows = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
    ).fetchall()
    return [str(r[0]) for r in rows]


def _sqlite_columns(conn: sqlite3.Connection, table: str) -> list[str]:
    rows = conn.execute(f'PRAGMA table_info("{table}")').fetchall()
    return [str(r[1]) for r in rows]


def _sqlite_rows(
    conn: sqlite3.Connection,
    *,
    table: str,
    limit: int,
    offset: int,
    order_by: str | None,
    desc: bool,
) -> dict[str, Any]:
    tables = set(_sqlite_tables(conn))
    if table not in tables:
        raise ValueError("tabla inválida")

    cols = _sqlite_columns(conn, table)
    if not cols:
        return {"columns": [], "rows": [], "limit": limit, "offset": offset}

    order_sql = ""
    if order_by:
        if order_by not in cols:
            raise ValueError("order_by inválido")
        order_sql = f' ORDER BY "{order_by}" ' + ("DESC" if desc else "ASC")

    q = f'SELECT * FROM "{table}"{order_sql} LIMIT ? OFFSET ?'
    rows = conn.execute(q, (int(limit), int(offset))).fetchall()
    out_rows = [list(r) for r in rows]
    return {"columns": cols, "rows": out_rows, "limit": int(limit), "offset": int(offset)}


_FLOAT_RE = re.compile(r"-?\d+(?:\.\d+)?")


def _parse_float(value: Any) -> float | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    m = _FLOAT_RE.search(s.replace(",", "."))
    if not m:
        return None
    try:
        return float(m.group(0))
    except Exception:
        return None


def _parse_epoch(value: Any) -> float | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None

    try:
        if s.endswith("Z"):
            s = s[:-1] + "+00:00"
        dt = datetime.fromisoformat(s)
        return float(dt.timestamp())
    except Exception:
        pass

    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            dt = datetime.strptime(s, fmt)
            return float(dt.timestamp())
        except Exception:
            continue
    return None


def _metrics() -> dict[str, Any]:
    """Calcula métricas desde las DB SQLite si existen."""

    dbs = {d["name"] for d in _list_sqlite_files()}

    providers: dict[str, Any] = {
        "growatt": {
            "device_count": 0,
            "current_voltage": None,
            "sparkline": [],
            "operativity_pct": None,
            "operativity_ok": 0,
            "operativity_total": 0,
            "operativity_started_at": None,
        },
        "shinemonitor": {
            "device_count": 0,
            "current_voltage": None,
            "sparkline": [],
            "operativity_pct": None,
            "operativity_ok": 0,
            "operativity_total": 0,
            "operativity_started_at": None,
        },
        "values": {
            "device_count": 0,
            "current_voltage": None,
            "sparkline": [],
            "operativity_pct": None,
            "operativity_ok": 0,
            "operativity_total": 0,
            "operativity_started_at": None,
        },
    }

    all_latest: list[float] = []

    # Growatt
    if "Voltage Growatt.sqlite" in dbs:
        try:
            conn = _sqlite_open_by_name("Voltage Growatt.sqlite")
            try:
                tables = [t for t in _sqlite_tables(conn) if t.startswith("g_")]
                providers["growatt"]["device_count"] = len(tables)

                started_at = _job_status("growatt").get("started_at")
                providers["growatt"]["operativity_started_at"] = started_at
                with _jobs_lock:
                    j = _jobs.get("growatt")
                    baseline = j.baseline_max_id if j else None

                if baseline is not None and tables:
                    ok = 0
                    total = len(tables)
                    for t in tables:
                        base_id = int((baseline.get(t) if isinstance(baseline, dict) else 0) or 0)
                        cur_id = _sqlite_max_id(conn, t)
                        if cur_id > base_id:
                            ok += 1
                    providers["growatt"]["operativity_ok"] = ok
                    providers["growatt"]["operativity_total"] = total
                    providers["growatt"]["operativity_pct"] = int(round((ok / total) * 100)) if total else None
                elif started_at and tables:
                    # Fallback (antiguo): por timestamp.
                    ok = 0
                    total = len(tables)
                    for t in tables:
                        try:
                            cols = set(_sqlite_columns(conn, t))
                            if "update_time" not in cols:
                                continue
                            row = conn.execute(
                                f'SELECT "update_time" FROM "{t}" ORDER BY id DESC LIMIT 1'
                            ).fetchone()
                            ts = _parse_epoch(row[0] if row else None)
                            if ts is not None and ts >= float(started_at):
                                ok += 1
                        except Exception:
                            continue
                    providers["growatt"]["operativity_ok"] = ok
                    providers["growatt"]["operativity_total"] = total
                    providers["growatt"]["operativity_pct"] = int(round((ok / total) * 100)) if total else None

                latest_vals: list[float] = []
                for t in tables:
                    try:
                        row = conn.execute(
                            f'SELECT "ac_output_voltage_frequency" FROM "{t}" ORDER BY id DESC LIMIT 1'
                        ).fetchone()
                        v = _parse_float(row[0] if row else None)
                        if v is not None:
                            latest_vals.append(v)
                    except Exception:
                        continue
                if latest_vals:
                    cur = sum(latest_vals) / len(latest_vals)
                    providers["growatt"]["current_voltage"] = cur
                    all_latest.extend(latest_vals)

                # Sparkline: usar primer monitor como referencia
                if tables:
                    t0 = tables[0]
                    try:
                        rows = conn.execute(
                            f'SELECT "ac_output_voltage_frequency" FROM "{t0}" ORDER BY id DESC LIMIT 30'
                        ).fetchall()
                        series = []
                        for r in reversed(rows):
                            vv = _parse_float(r[0])
                            if vv is not None:
                                series.append(vv)
                        providers["growatt"]["sparkline"] = series
                    except Exception:
                        pass
            finally:
                try:
                    conn.close()
                except Exception:
                    pass
        except Exception:
            pass

    # ShineMonitor
    if "Voltage  Shinemonitor.sqlite" in dbs:
        try:
            conn = _sqlite_open_by_name("Voltage  Shinemonitor.sqlite")
            try:
                try:
                    rows = conn.execute("SELECT table_name FROM meta_devices ORDER BY table_name").fetchall()
                    tables = [str(r[0]) for r in rows if r and r[0]]
                except Exception:
                    tables = []
                providers["shinemonitor"]["device_count"] = len(tables)

                st_sm = _job_status("shinemonitor")
                started_at = st_sm.get("started_at")
                providers["shinemonitor"]["operativity_started_at"] = started_at
                with _jobs_lock:
                    j = _jobs.get("shinemonitor")
                    baseline = j.baseline_max_id if j else None

                plants = _shinemonitor_plants(conn) if tables else {}

                # Por requerimiento: operatividad solo al final (no mientras corre).
                if (not st_sm.get("running")) and baseline is not None and tables:
                    if plants:
                        ok = 0
                        total = len(plants)
                        for _, ptables in plants.items():
                            plant_ok = False
                            for t in ptables:
                                base_id = int((baseline.get(t) if isinstance(baseline, dict) else 0) or 0)
                                cur_id = _sqlite_max_ok_voltage_id(conn, t, "Inverter Voltage(V)")
                                if cur_id > base_id:
                                    plant_ok = True
                                    break
                            if plant_ok:
                                ok += 1
                        providers["shinemonitor"]["operativity_ok"] = ok
                        providers["shinemonitor"]["operativity_total"] = total
                        providers["shinemonitor"]["operativity_pct"] = int(round((ok / total) * 100)) if total else None
                    else:
                        ok = 0
                        total = len(tables)
                        for t in tables:
                            base_id = int((baseline.get(t) if isinstance(baseline, dict) else 0) or 0)
                            cur_id = _sqlite_max_ok_voltage_id(conn, t, "Inverter Voltage(V)")
                            if cur_id > base_id:
                                ok += 1
                        providers["shinemonitor"]["operativity_ok"] = ok
                        providers["shinemonitor"]["operativity_total"] = total
                        providers["shinemonitor"]["operativity_pct"] = int(round((ok / total) * 100)) if total else None
                elif (not st_sm.get("running")) and started_at and tables:
                    try:
                        started = float(started_at)
                    except Exception:
                        started = None

                    if started is not None:
                        if plants:
                            ok = 0
                            total = len(plants)
                            for _, ptables in plants.items():
                                plant_ok = False
                                for t in ptables:
                                    try:
                                        cols = set(_sqlite_columns(conn, t))
                                        if "captured_at" not in cols:
                                            continue
                                        q = (
                                            f'SELECT "captured_at" FROM "{t}" '
                                            "WHERE status='OK' AND \"Inverter Voltage(V)\" IS NOT NULL AND \"Inverter Voltage(V)\" != '' "
                                            "ORDER BY captured_at DESC, id DESC LIMIT 1"
                                        )
                                        row = conn.execute(q).fetchone()
                                        ts = _parse_epoch(row[0] if row else None)
                                        if ts is not None and ts >= started:
                                            plant_ok = True
                                            break
                                    except Exception:
                                        continue
                                if plant_ok:
                                    ok += 1
                            providers["shinemonitor"]["operativity_ok"] = ok
                            providers["shinemonitor"]["operativity_total"] = total
                            providers["shinemonitor"]["operativity_pct"] = int(round((ok / total) * 100)) if total else None
                        else:
                            ok = 0
                            total = len(tables)
                            for t in tables:
                                try:
                                    cols = set(_sqlite_columns(conn, t))
                                    if "captured_at" not in cols:
                                        continue
                                    q = (
                                        f'SELECT "captured_at" FROM "{t}" '
                                        "WHERE status='OK' AND \"Inverter Voltage(V)\" IS NOT NULL AND \"Inverter Voltage(V)\" != '' "
                                        "ORDER BY captured_at DESC, id DESC LIMIT 1"
                                    )
                                    row = conn.execute(q).fetchone()
                                    ts = _parse_epoch(row[0] if row else None)
                                    if ts is not None and ts >= started:
                                        ok += 1
                                except Exception:
                                    continue
                            providers["shinemonitor"]["operativity_ok"] = ok
                            providers["shinemonitor"]["operativity_total"] = total
                            providers["shinemonitor"]["operativity_pct"] = int(round((ok / total) * 100)) if total else None

                latest_vals: list[float] = []
                for t in tables:
                    try:
                        row = conn.execute(
                            f'SELECT "Inverter Voltage(V)" FROM "{t}" '
                            "WHERE status='OK' AND \"Inverter Voltage(V)\" != '' "
                            "ORDER BY captured_at DESC, id DESC LIMIT 1"
                        ).fetchone()
                        v = _parse_float(row[0] if row else None)
                        if v is not None:
                            latest_vals.append(v)
                    except Exception:
                        continue
                if latest_vals:
                    cur = sum(latest_vals) / len(latest_vals)
                    providers["shinemonitor"]["current_voltage"] = cur
                    all_latest.extend(latest_vals)

                if tables:
                    t0 = tables[0]
                    try:
                        rows = conn.execute(
                            f'SELECT "Inverter Voltage(V)" FROM "{t0}" '
                            "WHERE \"Inverter Voltage(V)\" != '' "
                            "ORDER BY captured_at DESC, id DESC LIMIT 30"
                        ).fetchall()
                        series = []
                        for r in reversed(rows):
                            vv = _parse_float(r[0])
                            if vv is not None:
                                series.append(vv)
                        providers["shinemonitor"]["sparkline"] = series
                    except Exception:
                        pass
            finally:
                try:
                    conn.close()
                except Exception:
                    pass
        except Exception:
            pass

    # Values
    if "Voltage  Values.sqlite" in dbs:
        try:
            conn = _sqlite_open_by_name("Voltage  Values.sqlite")
            try:
                try:
                    rows = conn.execute("SELECT table_name FROM meta_monitors ORDER BY table_name").fetchall()
                    tables = [str(r[0]) for r in rows if r and r[0]]
                except Exception:
                    tables = []
                providers["values"]["device_count"] = len(tables)

                started_at = _job_status("values").get("started_at")
                providers["values"]["operativity_started_at"] = started_at
                with _jobs_lock:
                    j = _jobs.get("values")
                    baseline = j.baseline_max_id if j else None

                if baseline is not None and tables:
                    ok = 0
                    total = len(tables)
                    for t in tables:
                        base_id = int((baseline.get(t) if isinstance(baseline, dict) else 0) or 0)
                        cur_id = _sqlite_max_ok_id(conn, t)
                        if cur_id > base_id:
                            ok += 1
                    providers["values"]["operativity_ok"] = ok
                    providers["values"]["operativity_total"] = total
                    providers["values"]["operativity_pct"] = int(round((ok / total) * 100)) if total else None
                elif started_at and tables:
                    ok = 0
                    total = len(tables)
                    for t in tables:
                        try:
                            cols = set(_sqlite_columns(conn, t))
                            if "captured_at" not in cols:
                                continue
                            row = conn.execute(
                                f'SELECT "captured_at" FROM "{t}" ORDER BY captured_at DESC, id DESC LIMIT 1'
                            ).fetchone()
                            ts = _parse_epoch(row[0] if row else None)
                            if ts is not None and ts >= float(started_at):
                                ok += 1
                        except Exception:
                            continue
                    providers["values"]["operativity_ok"] = ok
                    providers["values"]["operativity_total"] = total
                    providers["values"]["operativity_pct"] = int(round((ok / total) * 100)) if total else None

                col = "Inverter_Voltage_V"
                latest_vals: list[float] = []
                for t in tables:
                    try:
                        cols = set(_sqlite_columns(conn, t))
                        use_col = col if col in cols else None
                        if not use_col:
                            continue
                        row = conn.execute(
                            f'SELECT "{use_col}" FROM "{t}" '
                            f'WHERE "{use_col}" IS NOT NULL AND "{use_col}" != \'\' '
                            "ORDER BY captured_at DESC, id DESC LIMIT 1"
                        ).fetchone()
                        v = _parse_float(row[0] if row else None)
                        if v is not None:
                            latest_vals.append(v)
                    except Exception:
                        continue
                if latest_vals:
                    cur = sum(latest_vals) / len(latest_vals)
                    providers["values"]["current_voltage"] = cur
                    all_latest.extend(latest_vals)

                if tables:
                    t0 = tables[0]
                    try:
                        cols0 = set(_sqlite_columns(conn, t0))
                        if col in cols0:
                            rows = conn.execute(
                                f'SELECT "{col}" FROM "{t0}" ORDER BY captured_at DESC, id DESC LIMIT 30'
                            ).fetchall()
                            series = []
                            for r in reversed(rows):
                                vv = _parse_float(r[0])
                                if vv is not None:
                                    series.append(vv)
                            providers["values"]["sparkline"] = series
                    except Exception:
                        pass
            finally:
                try:
                    conn.close()
                except Exception:
                    pass
        except Exception:
            pass

    avg = (sum(all_latest) / len(all_latest)) if all_latest else None
    total_devices = int(
        (providers["growatt"].get("device_count") or 0)
        + (providers["shinemonitor"].get("device_count") or 0)
        + (providers["values"].get("device_count") or 0)
    )

    return {
        "total_devices": total_devices,
        "avg_voltage": avg,
        "providers": providers,
        "sqlite_files": sorted(list(dbs)),
        "captured_at": time.time(),
    }


def _last_status_grid_file(provider: str) -> Path:
    return STORAGE_DIR / f"last-status-grid-{provider}.json"


def _save_last_status_grid(provider: str, payload: dict[str, Any]) -> None:
    try:
        _ensure_storage()
        # Escritura atómica: evita JSON truncado si se lee mientras se escribe.
        path = _last_status_grid_file(provider)
        tmp = path.with_suffix(path.suffix + ".tmp")
        tmp.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
        tmp.replace(path)
    except Exception:
        return


def _load_last_status_grid(provider: str) -> dict[str, Any] | None:
    try:
        path = _last_status_grid_file(provider)
        if not path.exists():
            return None
        data = json.loads(path.read_text(encoding="utf-8", errors="ignore") or "{}")
        if not isinstance(data, dict):
            return None
        return data
    except Exception:
        return None


def _normalize_cached_grid(
    *,
    provider: str,
    targets: list[str],
    cached: dict[str, Any] | None,
    running: bool,
) -> dict[str, Any]:
    """Alinea el grid guardado con los targets detectados actualmente.

    Evita padding fijo (p.ej. 77) y muestra solo plantas/monitores reales.
    """

    status_map: dict[str, str] = {}
    if cached and isinstance(cached.get("plants"), list):
        for it in cached.get("plants") or []:
            try:
                pid = str(it.get("plant_id") or "").strip()
                st = str(it.get("status") or "").strip() or "pending"
                if pid:
                    status_map[pid] = st
            except Exception:
                continue

    name_map = _grid_target_names_from_storage(provider)
    items = [
        {"plant_id": t, "name": (name_map.get(t) or ""), "status": status_map.get(t, "pending")}
        for t in targets
    ]
    ok = sum(1 for p in items if p["status"] == "ok")
    fail = sum(1 for p in items if p["status"] == "fail")
    retry = sum(1 for p in items if p["status"] == "retry")
    pending = sum(1 for p in items if p["status"] == "pending")
    return {
        "provider": provider,
        "running": bool(running),
        "plants": items,
        "ok": int(ok),
        "fail": int(fail),
        "retry": int(retry),
        "pending": int(pending),
        "total": int(len(items)),
        "updated_at": time.time(),
    }


def _status_grid(provider: str) -> dict[str, Any]:
    st = _job_status(provider)
    with _jobs_lock:
        job = _jobs.get(provider)
        baseline = job.baseline_max_id if job else None

    cached = _load_last_status_grid(provider)

    # Targets preferidos (para pintar grilla aunque no exista SQLite).
    # IMPORTANTE: al terminar una corrida, el scraper puede estar escribiendo el
    # snapshot (storage/*.json). Si ese JSON queda momentáneamente inválido,
    # _grid_targets_from_storage devuelve [] y la grilla caería a SQLite (menos celdas).
    # Por eso usamos fallback a job.grid_targets (si existe) y luego al cache.
    targets = _grid_targets_from_storage(provider)
    if (not targets) and job and job.grid_targets:
        targets = list(job.grid_targets)
    if (not targets) and cached and isinstance(cached.get("plants"), list):
        from_cache: list[str] = []
        for it in cached.get("plants") or []:
            try:
                pid = str((it or {}).get("plant_id") or "").strip()
            except Exception:
                pid = ""
            if pid:
                from_cache.append(pid)
        if from_cache:
            targets = from_cache

    name_map = _grid_target_names_from_storage(provider)
    if (not name_map) and cached and isinstance(cached.get("plants"), list):
        for it in cached.get("plants") or []:
            try:
                pid = str((it or {}).get("plant_id") or "").strip()
                nm = str((it or {}).get("name") or "").strip()
            except Exception:
                pid, nm = "", ""
            if pid and nm:
                name_map[pid] = nm

    # En ejecución: actualizar en vivo desde el log.
    if st.get("running") and job:
        if not job.grid_targets and targets:
            job.grid_targets = targets
            for t in targets:
                job.grid_status.setdefault(t, "pending")
        _update_job_live_grid(provider, job)
        tlist = job.grid_targets or targets
        items = [
            {"plant_id": t, "name": (name_map.get(t) or ""), "status": (job.grid_status.get(t) or "pending")}
            for t in (tlist or [])
        ]
        ok = sum(1 for p in items if p["status"] == "ok")
        fail = sum(1 for p in items if p["status"] == "fail")
        retry = sum(1 for p in items if p["status"] == "retry")
        pending = sum(1 for p in items if p["status"] == "pending")
        return {
            "provider": provider,
            "running": True,
            "plants": items,
            "ok": int(ok),
            "fail": int(fail),
            "retry": int(retry),
            "pending": int(pending),
            "total": int(len(items)),
            "updated_at": time.time(),
        }

    # No está corriendo:
    # - Growatt/Values: mostramos el último grid guardado (derivado del log) y lo normalizamos.
    # - ShineMonitor: intentamos recalcular desde SQLite usando baseline; si no se puede, caemos al cache.
    if provider in ("growatt", "values"):
        if targets:
            return _normalize_cached_grid(provider=provider, targets=targets, cached=cached, running=False)
        return cached if cached else {
            "provider": provider,
            "running": False,
            "plants": [],
            "ok": 0,
            "fail": 0,
            "retry": 0,
            "pending": 0,
            "total": 0,
            "updated_at": time.time(),
        }

    db_name = _provider_db_name(provider)
    dbs = {d["name"] for d in _list_sqlite_files()}
    if not db_name or db_name not in dbs:
        return cached if cached else {
            "provider": provider,
            "running": False,
            "plants": [],
            "ok": 0,
            "fail": 0,
            "retry": 0,
            "pending": 0,
            "total": 0,
            "updated_at": time.time(),
        }

    statuses: dict[str, str] = {}
    items: list[dict[str, Any]] = []

    # ShineMonitor: si el servidor reinició y no hay baseline, aun así podemos
    # mostrar un estado consistente leyendo la última fila real en SQLite.

    # Si no hay targets desde storage, caer a targets desde SQLite.
    if not targets:
        try:
            conn = _sqlite_open_by_name(db_name)
            try:
                if provider == "shinemonitor":
                    targets = sorted(
                        _shinemonitor_plants(conn).keys(),
                        key=lambda s: int(s) if str(s).isdigit() else str(s),
                    )
                else:
                    targets = _provider_expected_tables(conn, provider)
            finally:
                conn.close()
        except Exception:
            targets = []

    if provider == "shinemonitor":
        # Targets por planta
        plants_map: dict[str, list[str]] = {}
        try:
            conn = _sqlite_open_by_name(db_name)
            try:
                plants_map = _shinemonitor_plants(conn)
            finally:
                conn.close()
        except Exception:
            plants_map = {}

        # IMPORTANTE: el conteo/orden de la grilla debe reflejar lo que se está procesando
        # (snapshot en storage). meta_devices puede no incluir plantas sin tablas (NO_INVERTER/NO_DEVICES).
        plant_ids = list(targets) if targets else sorted(
            plants_map.keys(), key=lambda s: int(s) if str(s).isdigit() else str(s)
        )
        for pid in plant_ids:
            statuses[pid] = "pending"

        # Fail incremental durante ejecución; al finalizar, parseo completo.
        if job and st.get("running"):
            for pid in _parse_shinemonitor_fail_plants_incremental(job):
                statuses[pid] = "fail"
        if job and (not st.get("running")):
            for pid in _parse_shinemonitor_fail_plants_full(job.log_path):
                statuses[pid] = "fail"

        # Caso 1: hay baseline => operatividad por inserciones OK desde baseline.
        if baseline is not None and plants_map:
            try:
                conn = _sqlite_open_by_name(db_name)
                try:
                    for pid in plant_ids:
                        if statuses.get(pid) == "fail":
                            continue
                        tables = plants_map.get(pid) or []
                        if not tables:
                            # Sin tablas (no hubo devices) => decidir por el último evento si existe.
                            ev = _shinemonitor_latest_event_status(conn=conn, plant_id=pid)
                            if ev and ev != "OK":
                                statuses[pid] = "fail"
                            elif not st.get("running"):
                                statuses[pid] = "fail"
                            continue
                        plant_ok = False
                        for t in tables:
                            base_id = int((baseline.get(t) if isinstance(baseline, dict) else 0) or 0)
                            cur_id = _sqlite_max_ok_voltage_id(conn, t, "Inverter Voltage(V)")
                            if cur_id > base_id:
                                plant_ok = True
                                break
                        if plant_ok:
                            statuses[pid] = "ok"
                        else:
                            # Al terminar: si no insertó nada OK y no fue NO_TABLE/NO_TAB, es FAIL.
                            if not st.get("running"):
                                statuses[pid] = "fail"
                finally:
                    conn.close()
            except Exception:
                pass

        # Caso 2: sin baseline (reinicio del servidor, o status histórico):
        # derivar status desde la última fila real de cada device en SQLite.
        if baseline is None and (not st.get("running")):
            try:
                conn = _sqlite_open_by_name(db_name)
                try:
                    for pid in plant_ids:
                        tables = plants_map.get(pid) or []
                        if tables:
                            statuses[pid] = _shinemonitor_plant_status_from_db(
                                conn=conn,
                                plant_id=pid,
                                plants_map=plants_map,
                            )
                        else:
                            # Sin tablas: usar último evento si existe, si no pending.
                            ev = _shinemonitor_latest_event_status(conn=conn, plant_id=pid)
                            if ev and ev != "OK":
                                statuses[pid] = "fail"
                finally:
                    conn.close()
            except Exception:
                pass

        items = [
            {"plant_id": pid, "name": (name_map.get(pid) or ""), "status": statuses.get(pid, "pending")}
            for pid in plant_ids
        ]
    else:
        # Targets por monitor/tabla
        tables: list[str] = []
        try:
            conn = _sqlite_open_by_name(db_name)
            try:
                tables = _provider_expected_tables(conn, provider)
            finally:
                conn.close()
        except Exception:
            tables = []

        tables = [t for t in tables if t]
        for t in tables:
            statuses[t] = "pending"

        if baseline is not None and tables:
            try:
                conn = _sqlite_open_by_name(db_name)
                try:
                    for t in tables:
                        base_id = int((baseline.get(t) if isinstance(baseline, dict) else 0) or 0)
                        if provider == "values":
                            cur_id = _sqlite_max_ok_voltage_id(conn, t, "Inverter_Voltage_V")
                        else:
                            cur_id = _sqlite_max_id(conn, t)
                        statuses[t] = "ok" if cur_id > base_id else ("pending" if st.get("running") else "fail")
                finally:
                    conn.close()
            except Exception:
                pass

        items = [{"plant_id": t, "name": (name_map.get(t) or ""), "status": statuses.get(t, "pending")} for t in tables]

    ok = sum(1 for p in items if p["status"] == "ok")
    fail = sum(1 for p in items if p["status"] == "fail")
    retry = sum(1 for p in items if p["status"] == "retry")
    pending = sum(1 for p in items if p["status"] == "pending")
    payload = {
        "provider": provider,
        "running": bool(st.get("running")),
        "plants": items,
        "ok": int(ok),
        "fail": int(fail),
        "retry": int(retry),
        "pending": int(pending),
        "total": int(len(items)),
        "updated_at": time.time(),
    }
    if not st.get("running"):
        _save_last_status_grid(provider, payload)
    return payload


def _finalize_run(provider: str, job: Job) -> None:
    """Al terminar una corrida, calcula el grid final y guarda eventos una sola vez."""

    try:
        # Para Growatt/Values: el grid final se basa en inserciones (log) y reintentos.
        if provider in ("growatt", "values"):
            # Asegurar que el parseo alcanzó el final del archivo.
            _update_job_live_grid(provider, job)
            targets = job.grid_targets or _grid_targets_from_storage(provider)
            # Cerrar último target si quedó en retry.
            if job.grid_current and (job.grid_status.get(job.grid_current) == "retry"):
                job.grid_status[job.grid_current] = "fail"
            # Convertir pending/retry a fail en cierre.
            for t in targets:
                stt = (job.grid_status.get(t) or "pending").strip()
                if stt in ("pending", "retry"):
                    job.grid_status[t] = "fail"
            items = [{"plant_id": t, "status": (job.grid_status.get(t) or "fail")} for t in targets]
            ok = sum(1 for p in items if p["status"] == "ok")
            fail = sum(1 for p in items if p["status"] == "fail")
            retry = sum(1 for p in items if p["status"] == "retry")
            pending = sum(1 for p in items if p["status"] == "pending")
            grid = {
                "provider": provider,
                "running": False,
                "plants": items,
                "ok": int(ok),
                "fail": int(fail),
                "retry": int(retry),
                "pending": int(pending),
                "total": int(len(items)),
                "updated_at": time.time(),
            }
        else:
            grid = _status_grid(provider)

        # Persistir explícitamente como "último scrapeo".
        _save_last_status_grid(provider, grid)

        # Emitir eventos ok/fail para la serie Plantas/Hora.
        ts = time.time()
        for item in grid.get("plants") or []:
            pid = str(item.get("plant_id") or "").strip()
            st = str(item.get("status") or "").strip()
            if not pid:
                continue
            if st == "ok":
                key = f"ok:{pid}:{job.started_at}"
                if key in job.seen_plant_events:
                    continue
                job.seen_plant_events.add(key)
                _append_plant_event(provider=provider, plant_id=pid, outcome="ok", ts=ts)
            elif st == "fail":
                key = f"fail:{pid}:{job.started_at}"
                if key in job.seen_plant_events:
                    continue
                job.seen_plant_events.add(key)
                _append_plant_event(provider=provider, plant_id=pid, outcome="fail", ts=ts)
    except Exception:
        return


def _basic_auth_expected() -> str | None:
    """Devuelve el token esperado "user:pass" o None si no hay auth configurada.

    Para habilitar Basic Auth, define:
      - WEBUI_BASIC_AUTH="usuario:password"
    o alternativamente:
      - WEBUI_USER="usuario" y WEBUI_PASS="password"
    """

    raw = (os.environ.get("WEBUI_BASIC_AUTH") or "").strip()
    if raw:
        return raw
    user = (os.environ.get("WEBUI_USER") or "").strip()
    pwd = (os.environ.get("WEBUI_PASS") or "").strip()
    if user and pwd:
        return f"{user}:{pwd}"
    return None


def _basic_auth_check(handler: BaseHTTPRequestHandler) -> bool:
    expected = _basic_auth_expected()
    if not expected:
        return True

    header = (handler.headers.get("Authorization") or "").strip()
    if not header.lower().startswith("basic "):
        return False

    token = header.split(" ", 1)[1].strip()
    try:
        decoded = base64.b64decode(token.encode("ascii"), validate=True).decode("utf-8")
    except Exception:
        return False
    return hmac.compare_digest(decoded, expected)


def _basic_auth_deny(handler: BaseHTTPRequestHandler) -> None:
    handler.send_response(401)
    handler.send_header("WWW-Authenticate", 'Basic realm="Voltguard"')
    handler.send_header("Content-Type", "text/plain; charset=utf-8")
    handler.end_headers()
    handler.wfile.write(b"Unauthorized")


class Handler(BaseHTTPRequestHandler):
    def log_message(self, format: str, *args: Any) -> None:
        # Silenciar log ruidoso del servidor
        return

    def do_GET(self) -> None:
        if not _basic_auth_check(self):
            _basic_auth_deny(self)
            return

        # Asegurar reset diario del reporte.
        _ensure_report_reset_thread()

        parsed = urlparse(self.path)
        path = parsed.path

        # Rutas multipágina (mismo bundle JS/CSS)
        routes: dict[str, str] = {
            "/": "dashboard.html",
            "/index.html": "dashboard.html",
            "/dashboard": "dashboard.html",
            "/providers": "providers.html",
            "/reports": "reports.html",
            "/sqlite": "sqlite.html",
            "/logs": "logs.html",
            "/logs.html": "logs.html",
            "/settings": "settings.html",
            "/tunnel": "tunnel.html",
        }

        if path in routes:
            self._serve_file(WEB_DIR / routes[path], content_type="text/html; charset=utf-8")
            return

        if path == "/" or path == "/index.html":
            self._serve_file(WEB_DIR / "index.html", content_type="text/html; charset=utf-8")
            return

        if path == "/app.js":
            self._serve_file(WEB_DIR / "app.js", content_type="text/javascript; charset=utf-8")
            return

        if path == "/styles.css":
            self._serve_file(WEB_DIR / "styles.css", content_type="text/css; charset=utf-8")
            return

        if path in ("/rayo.ico", "/favicon.ico"):
            self._serve_file(WEB_DIR / "rayo.ico", content_type="image/x-icon")
            return

        if path == "/api/config":
            keys = [
                "HEADLESS",
                "BROWSER",
                "SHINE_DEFAULT_TIMEOUT_MS",
                "SHINE_NAV_TIMEOUT_MS",
                "VALUES_TURBO",
                "VALUES_DEFAULT_TIMEOUT_MS",
                "VALUES_NAV_TIMEOUT_MS",
                "VALUES_USE_DEVICE_LIST",
                "VALUES_LIMIT_MONITORS",
                "WEBUI_HOST",
                "WEBUI_PORT",
            ]
            env_summary = {
                k: {"value": os.environ.get(k), "defined_in_dotenv": _dotenv_defines(k)} for k in keys
            }
            _json_response(
                self,
                200,
                {
                    "auth_enabled": _basic_auth_expected() is not None,
                    "host": os.environ.get("WEBUI_HOST", "0.0.0.0"),
                    "port": int(os.environ.get("WEBUI_PORT", "8000")),
                    "local_ip": _get_local_ip(),
                    "local_url": f"http://{_get_local_ip()}:{int(os.environ.get('WEBUI_PORT', '8000'))}/",
                    "sqlite_files": _list_sqlite_files(),
                    "env": env_summary,
                    "providers": [
                        {"key": k, "label": v["label"], "script": v["script"], "log": v["log"]}
                        for k, v in PROVIDERS.items()
                    ],
                },
            )
            return

        if path == "/api/metrics":
            _json_response(self, 200, _metrics())
            return

        if path == "/api/providers/monitors":
            result: dict[str, Any] = {}
            # ShineMonitor
            sm_path = STORAGE_DIR / "shinemonitor-plants.json"
            if sm_path.exists():
                try:
                    sm_data = json.loads(sm_path.read_text(encoding="utf-8", errors="ignore") or "{}")
                    sm_plants = sm_data.get("plants") or []
                    result["shinemonitor"] = {
                        "label": "ShineMonitor",
                        "count": len(sm_plants),
                        "captured_at": sm_data.get("captured_at", ""),
                        "items": [{"id": p.get("plant_id", ""), "name": p.get("name", "")} for p in sm_plants],
                    }
                except Exception:
                    result["shinemonitor"] = {"label": "ShineMonitor", "count": 0, "items": [], "error": "No se pudo leer el JSON"}
            else:
                result["shinemonitor"] = {"label": "ShineMonitor", "count": 0, "items": [], "error": "Archivo no encontrado"}

            # Growatt
            gw_path = STORAGE_DIR / "growatt-plants.json"
            if gw_path.exists():
                try:
                    gw_data = json.loads(gw_path.read_text(encoding="utf-8", errors="ignore") or "{}")
                    gw_plants = gw_data.get("dropdown_plants") or gw_data.get("home_plants") or []
                    result["growatt"] = {
                        "label": "Growatt",
                        "count": len(gw_plants),
                        "captured_at": gw_data.get("generated_at", ""),
                        "items": [{"id": str(p.get("id") or p.get("index", "")), "name": p.get("name", "")} for p in gw_plants],
                    }
                except Exception:
                    result["growatt"] = {"label": "Growatt", "count": 0, "items": [], "error": "No se pudo leer el JSON"}
            else:
                result["growatt"] = {"label": "Growatt", "count": 0, "items": [], "error": "Archivo no encontrado"}

            # Values
            vals_path = STORAGE_DIR / "values-monitors.json"
            if vals_path.exists():
                try:
                    vals_data = json.loads(vals_path.read_text(encoding="utf-8", errors="ignore") or "{}")
                    vals_monitors = vals_data.get("monitors") or []
                    result["values"] = {
                        "label": "Values",
                        "count": len(vals_monitors),
                        "captured_at": vals_data.get("captured_at", ""),
                        "items": [{"id": m.get("external_id", ""), "name": m.get("name", "")} for m in vals_monitors],
                    }
                except Exception:
                    result["values"] = {"label": "Values", "count": 0, "items": [], "error": "No se pudo leer el JSON"}
            else:
                result["values"] = {"label": "Values", "count": 0, "items": [], "error": "Archivo no encontrado"}

            _json_response(self, 200, result)
            return

        if path == "/api/report/status":
            _json_response(self, 200, _report_status())
            return

        if path == "/api/report/history":
            _json_response(self, 200, {"files": _report_history_list()})
            return

        if path == "/api/report/download":
            qs = parse_qs(parsed.query)
            name = (qs.get("file") or [""])[0]
            p = _safe_report_path(name)
            if not p or not p.exists() or not p.is_file():
                _json_response(self, 404, {"error": "archivo no encontrado"})
                return
            try:
                data = p.read_bytes()
                self.send_response(200)
                self.send_header(
                    "Content-Type",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                self.send_header(
                    "Content-Disposition",
                    f'attachment; filename="{p.name}"',
                )
                self.send_header("Content-Length", str(len(data)))
                self.end_headers()
                self.wfile.write(data)
            except Exception as e:
                _json_response(self, 400, {"error": f"{type(e).__name__}: {e}"})
            return

        if path == "/api/report/export":
            qs = parse_qs(parsed.query)
            slot = (qs.get("slot") or [""])[0]
            rid = (qs.get("rid") or [""])[0]
            try:
                with _report_lock:
                    # Si el navegador/túnel repite el mismo request (mismo rid),
                    # no generar múltiples copias en historial.
                    duplicate = bool(rid and _export_rid_seen(rid))
                    if not duplicate:
                        out_path = _generate_or_update_report(slot=slot)
                        _report_history_copy_from_current(slot=slot)
                        _export_rid_mark(rid)
                    else:
                        out_path = REPORT_PATH
                        if not out_path.exists():
                            out_path = _generate_or_update_report(slot=slot)
                slot_c = _canonical_slot(slot)
                slot_names_map = {
                    "manana": "Mañana",
                    "mediodia": "Medio Dia",
                    "tarde": "Tarde",
                    "medianoche": "MediaNoche",
                }
                slot_name = slot_names_map.get(slot_c, "Reporte")
                dl_filename = f"Reporte de Monitores_{slot_name}.xlsx"
                from urllib.parse import quote
                safe_filename = quote(dl_filename)

                data = out_path.read_bytes()
                self.send_response(200)
                self.send_header(
                    "Content-Type",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                self.send_header(
                    "Content-Disposition",
                    f"attachment; filename*=UTF-8''{safe_filename}",
                )
                self.send_header("Content-Length", str(len(data)))
                self.end_headers()
                self.wfile.write(data)
            except Exception as e:
                _json_response(self, 400, {"error": f"{type(e).__name__}: {e}"})
            return

        if path == "/api/status-grid":
            qs = parse_qs(parsed.query)
            provider = (qs.get("provider") or ["shinemonitor"])[0].strip() or "shinemonitor"
            if provider not in ("growatt", "shinemonitor", "values"):
                _json_response(self, 400, {"error": "provider inválido"})
                return
            _json_response(self, 200, _status_grid(provider))
            return

        if path == "/api/network-load":
            qs = parse_qs(parsed.query)
            provider = (qs.get("provider") or ["shinemonitor"])[0].strip() or "shinemonitor"
            if provider not in ("growatt", "shinemonitor", "values"):
                _json_response(self, 400, {"error": "provider inválido"})
                return
            _json_response(self, 200, _network_load_series(provider))
            return

        if path == "/api/sqlite/files":
            _json_response(self, 200, {"files": _list_sqlite_files()})
            return

        if path == "/api/sqlite/tables":
            qs = parse_qs(parsed.query)
            db = (qs.get("db") or [""])[0]
            if not db:
                _json_response(self, 400, {"error": "db requerido"})
                return
            try:
                conn = _sqlite_open_by_name(db)
                try:
                    _json_response(self, 200, {"db": db, "tables": _sqlite_tables(conn)})
                finally:
                    try:
                        conn.close()
                    except Exception:
                        pass
            except Exception as e:
                _json_response(self, 400, {"error": f"{type(e).__name__}: {e}"})
            return

        if path == "/api/sqlite/rows":
            qs = parse_qs(parsed.query)
            db = (qs.get("db") or [""])[0]
            table = (qs.get("table") or [""])[0]
            order_by = (qs.get("order_by") or [""])[0].strip() or None
            desc = ((qs.get("desc") or ["1"])[0].strip() != "0")
            try:
                limit = int((qs.get("limit") or ["100"])[0])
            except Exception:
                limit = 100
            try:
                offset = int((qs.get("offset") or ["0"])[0])
            except Exception:
                offset = 0
            limit = max(1, min(500, limit))
            offset = max(0, offset)

            if not db or not table:
                _json_response(self, 400, {"error": "db y table requeridos"})
                return
            try:
                conn = _sqlite_open_by_name(db)
                try:
                    data = _sqlite_rows(
                        conn,
                        table=table,
                        limit=limit,
                        offset=offset,
                        order_by=order_by,
                        desc=bool(desc),
                    )
                finally:
                    try:
                        conn.close()
                    except Exception:
                        pass
                data.update({"db": db, "table": table, "order_by": order_by, "desc": bool(desc)})
                _json_response(self, 200, data)
            except Exception as e:
                _json_response(self, 400, {"error": f"{type(e).__name__}: {e}"})
            return

        if path == "/api/providers":
            _json_response(
                self,
                200,
                {
                    "providers": [
                        {"key": k, "label": v["label"], "script": v["script"], "log": v["log"]}
                        for k, v in PROVIDERS.items()
                    ]
                },
            )
            return

        if path == "/api/status":
            qs = parse_qs(parsed.query)
            provider = (qs.get("provider") or [""])[0]
            if provider:
                _json_response(self, 200, _job_status(provider))
            else:
                _json_response(self, 200, {k: _job_status(k) for k in PROVIDERS})
            return

        if path == "/api/log":
            qs = parse_qs(parsed.query)
            provider = (qs.get("provider") or [""])[0]
            pos_raw = (qs.get("pos") or ["0"])[0]
            try:
                pos = int(pos_raw)
            except Exception:
                pos = 0
            if not provider:
                _json_response(self, 400, {"error": "provider requerido"})
                return
            _json_response(self, 200, _read_log(provider, pos))
            return

        self.send_response(404)
        self.end_headers()

    def do_POST(self) -> None:
        if not _basic_auth_check(self):
            _basic_auth_deny(self)
            return

        _ensure_report_reset_thread()
        parsed = urlparse(self.path)
        if parsed.path == "/api/run":
            body = _read_body_json(self)
            provider = str(body.get("provider") or "").strip().lower()
            if provider not in PROVIDERS:
                _json_response(self, 400, {"error": "provider inválido"})
                return

            status = _job_status(provider)
            if status.get("running"):
                _json_response(self, 409, {"error": "ya está ejecutándose", "status": status})
                return

            try:
                job = _start_job(provider)
            except Exception as e:
                _json_response(self, 500, {"error": f"no se pudo iniciar: {type(e).__name__}: {e}"})
                return

            _json_response(
                self,
                200,
                {
                    "ok": True,
                    "status": _job_status(job.provider),
                    "others_running": _other_running_providers(current=provider),
                },
            )
            return

        if parsed.path == "/api/settings/save":
            body = _read_body_json(self)
            env_path = BASE_DIR / ".env"
            
            try:
                # Mantenemos las líneas y comentarios existentes
                lines = []
                if env_path.exists():
                    with open(env_path, "r", encoding="utf-8") as f:
                        lines = f.readlines()
                
                updated_keys = set()
                new_lines = []
                for line in lines:
                    stripped = line.strip()
                    if stripped and not stripped.startswith("#") and "=" in stripped:
                        key = stripped.split("=", 1)[0].strip()
                        if key in body:
                            new_lines.append(f"{key}={body[key]}\n")
                            updated_keys.add(key)
                        else:
                            new_lines.append(line)
                    else:
                        new_lines.append(line)
                
                # Añadir claves nuevas al final
                for k, v in body.items():
                    if k not in updated_keys:
                        # Si no terminó con newline, agregar una antes de las nuevas keys
                        if new_lines and not new_lines[-1].endswith("\n"):
                            new_lines[-1] += "\n"
                        new_lines.append(f"{k}={v}\n")
                
                with open(env_path, "w", encoding="utf-8") as f:
                    f.writelines(new_lines)
                
                _json_response(self, 200, {"ok": True})
            except Exception as e:
                _json_response(self, 500, {"error": str(e)})
            return

        if parsed.path == "/api/report/clear":
            try:
                with _report_lock:
                    _clear_report()
                _json_response(self, 200, {"ok": True})
            except Exception as e:
                _json_response(self, 400, {"error": f"{type(e).__name__}: {e}"})
            return

        if parsed.path == "/api/report/delete-all":
            try:
                _ensure_reports_dir()
                files_deleted = 0
                for p in REPORTS_DIR.glob("*.xlsx"):
                    if p.name == REPORT_FILENAME:
                        continue
                    try:
                        p.unlink(missing_ok=True)
                        files_deleted += 1
                    except Exception:
                        continue
                _json_response(self, 200, {"ok": True, "deleted": files_deleted})
            except Exception as e:
                _json_response(self, 400, {"error": f"{type(e).__name__}: {e}"})
            return

        if parsed.path == "/api/report/delete":
            body = _read_body_json(self)
            name = str(body.get("file") or "").strip()
            p = _safe_report_path(name)
            if not p or not p.exists() or not p.is_file():
                _json_response(self, 404, {"error": "archivo no encontrado"})
                return
            try:
                # Borrar archivo del historial
                p.unlink(missing_ok=True)
                _json_response(self, 200, {"ok": True})
            except Exception as e:
                _json_response(self, 400, {"error": f"{type(e).__name__}: {e}"})
            return

        if parsed.path == "/api/clear":
            body = _read_body_json(self)
            provider = str(body.get("provider") or "").strip().lower()
            if provider not in PROVIDERS:
                _json_response(self, 400, {"error": "provider inválido"})
                return
            _json_response(self, 200, _clear_log(provider))
            return

        if parsed.path == "/api/stop":
            body = _read_body_json(self)
            provider = str(body.get("provider") or "").strip().lower()
            if provider not in PROVIDERS:
                _json_response(self, 400, {"error": "provider inválido"})
                return
            _json_response(self, 200, _stop_job(provider))
            return
        if parsed.path == "/api/sqlite/drop-tables":
            body = _read_body_json(self)
            db_name = str(body.get("db") or "").strip()
            tables_to_drop = body.get("tables") or []
            if not db_name or not tables_to_drop:
                _json_response(self, 400, {"error": "Faltan parámetros: db y tables"})
                return
            if not isinstance(tables_to_drop, list):
                _json_response(self, 400, {"error": "tables debe ser una lista"})
                return
            try:
                conn_ro = _sqlite_open_by_name(db_name)
                try:
                    existing = {r[0] for r in conn_ro.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
                finally:
                    conn_ro.close()
            except Exception as e:
                _json_response(self, 400, {"error": f"No se pudo abrir DB: {e}"})
                return
            protected = {"meta_plants", "meta_devices", "meta_monitors", "meta_columns", "plant_events", "sqlite_sequence"}
            safe_tables = [t for t in tables_to_drop if t in existing and t not in protected]
            if not safe_tables:
                _json_response(self, 400, {"error": "No hay tablas válidas para eliminar"})
                return
            dropped: list[str] = []
            errors: list[str] = []
            try:
                conn = _sqlite_open_rw_by_name(db_name)
                try:
                    for tbl in safe_tables:
                        try:
                            conn.execute(f'DROP TABLE IF EXISTS "{tbl}"')
                            for meta in ("meta_monitors", "meta_devices", "meta_columns"):
                                try:
                                    conn.execute(f'DELETE FROM "{meta}" WHERE table_name=?', (tbl,))
                                except Exception:
                                    pass
                            dropped.append(tbl)
                        except Exception as e:
                            errors.append(f"{tbl}: {e}")
                    conn.commit()
                finally:
                    conn.close()
            except Exception as e:
                _json_response(self, 500, {"error": f"Error al abrir DB para escritura: {e}"})
                return
            _json_response(self, 200, {
                "ok": True,
                "dropped": dropped,
                "errors": errors,
                "message": f"Eliminadas {len(dropped)} tabla(s)" + (f" con {len(errors)} error(es)" if errors else ""),
            })
            return

        self.send_response(404)
        self.end_headers()

    def _serve_file(self, path: Path, *, content_type: str) -> None:
        if not path.exists():
            self.send_response(404)
            self.end_headers()
            return
        data = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)


def _get_local_ip() -> str:
    import socket
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"


def main() -> None:
    host = os.environ.get("WEBUI_HOST", "0.0.0.0")
    port = int(os.environ.get("WEBUI_PORT", "8000"))

    if not WEB_DIR.exists():
        raise SystemExit("No existe carpeta webui/")

    server = ThreadingHTTPServer((host, port), Handler)
    local_ip = _get_local_ip()
    print(f"Web UI (Local): http://localhost:{port}/")
    if local_ip != "127.0.0.1":
        print(f"Web UI (Red Local): http://{local_ip}:{port}/")
    server.serve_forever()


if __name__ == "__main__":
    main()
